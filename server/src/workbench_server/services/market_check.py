"""The gate-closure / bid-window gate — M6's second *domain*
:class:`~workbench_server.services.validation.ValidationCheck`, and the one with no
prior art to copy.

Reconciliation asks *do the numbers agree*. This asks the question that comes first
for anything destined for a power exchange: **is this artifact a legal bid at all?**
A schedule whose numbers reconcile perfectly is worthless if it was stamped one
second after gate closure, if its delivery periods sit off the market time unit, or
if it carries 24 hourly rows for a day that has 25 of them.

Five violation classes, each its own :class:`EvidenceItem` so
:func:`~workbench_server.services.validation.derive_risk` can do the right thing
(a late bid is ``high`` even when the grid is spotless):

* **Gate closure is local wall-clock.** The Nordic day-ahead auction closes at 12:00
  *in Oslo* — 10:00 UTC in winter, 11:00 UTC in summer. Comparing a UTC-naive
  submission timestamp against that is the single bug this check exists to catch, so
  when a naive ``submitted_at`` is *before* the deadline read as local and *after* it
  read as UTC, the check does not pick a winner: it reports the disagreement and
  names the fix (put an offset on the timestamp). Submitted **at** gate closure is
  late — the auction is closed at 12:00:00.000.
* **Delivery periods must sit on the market's grid**, tested in *instants* rather
  than in wall clocks. Every period start is resolved to a UTC instant and measured
  against the start of its own local day, so the test stays correct across a
  transition instead of accidentally passing because ``:00`` looks aligned.
* **A local day is not 24 hours.** The expected period count comes from the zone's
  own UTC offsets — 25 hourly periods (100 quarter-hourly) on the Nordic fall-back
  Sunday, 23 (92) on the spring-forward one. A fall-back day that validates as 24/96
  is precisely the silent failure this refuses.
* **Ambiguous local times must be disambiguated.** On a fall-back day one wall clock
  occurs twice. A file whose timestamps carry no UTC offset and no fold column is
  read in order of appearance — the shipped
  :func:`~workbench_server.services.reconciliation.is_ambiguous_local` rule, which
  asks the *zone* whether that wall clock genuinely repeats rather than trusting row
  order — and the reading is reported as a ``warn`` naming both fixes. Its
  spring-forward twin, a wall clock that never existed, is a ``fail``.
* **Unit headers are checked against the period duration.** ``MW`` and ``MWh`` are
  numerically identical over a 60-minute MTU and differ by a factor of four over a
  15-minute one, which is exactly how a correct-looking file becomes 4x wrong the
  day the market switches. Prices reuse the reconciliation unit table, so a NOK
  column on a EUR market is the same **principled refusal** — never an invented FX
  rate, and never "unknown unit".

Fake-first, like its sibling: ``.xlsx`` through openpyxl and ``.csv`` through the
stdlib, no Office, no network and no live market data, so all three CI legs prove it.
Timezones are :mod:`zoneinfo` against the ``tzdata`` package the project already
pins — no new dependency.

**The rules are server-owned** (:data:`~workbench_server.models.market.MARKETS`),
selected by id. The artifact never states its own gate closure, because a rule the
subject supplies is a rule the subject can define away.
"""

from __future__ import annotations

import csv
import re
from dataclasses import dataclass, field
from datetime import UTC, date, datetime, time, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo

import openpyxl
import structlog
from pydantic import ValidationError

from workbench_server.models.market import (
    MARKETS,
    RESOLUTION_MINUTES,
    DayCoverage,
    GateClosureFinding,
    MarketCheckSpec,
    MarketComplianceReport,
    MarketSpec,
    PeriodViolation,
    UnitFinding,
    market_ids,
)
from workbench_server.models.validation import (
    CheckOutcome,
    EvidenceItem,
    EvidenceTruncation,
)
from workbench_server.services.reconciliation import (
    UNIT_TO_BASE,
    convert,
    is_ambiguous_local,
)
from workbench_server.services.validation import ValidationContext

log = structlog.get_logger()

#: The artifact formats the check reads. Both are read locally and deterministically:
#: ``.xlsx`` with openpyxl (already a dependency, for the reconciliation gate) and
#: ``.csv`` with the stdlib. Anything else is a ``fail`` that names these two.
SUPPORTED_SUFFIXES = (".xlsx", ".csv")

#: Cap on the data rows read out of one artifact. A leap year of quarter-hourly
#: periods is 35,136 rows, so this clears a year's schedule with headroom; past it
#: the check refuses and names the cap rather than chewing through a file that is
#: not a bid.
MAX_ROWS = 40_000

#: Cap on the violation rows carried inside one stored report (AXI shape 1).
MAX_VIOLATIONS = 200

#: Cap on the *named* bad-delivery-day evidence lines. A year-long schedule with a
#: systematic error would otherwise emit 365 lines and blow the result's budget;
#: past this they roll into one line that says how many were withheld.
MAX_DAY_LINES = 20

#: How many delivery days one *passing* line names before it rolls the rest up. A
#: year-long schedule that is entirely correct still owes the reader a sentence, not
#: 365 dates.
MAX_NAMED_DAYS = 5

#: How much of an unreadable cell is quoted back in the evidence.
MAX_QUOTE = 40


class _CheckBlocked(Exception):
    """Internal: a condition that stops the whole check with one stated reason.

    Surfaced as a single ``fail`` :class:`EvidenceItem`, never as an empty result.
    The frame's ``blocked`` risk is reserved for *no evidence at all*
    (``services/validation.py``), and returning nothing here would produce a result
    whose summary could not name the reason — so an unparseable artifact is a
    stated ``fail``, which is also what ``ReconciliationCheck`` does. Either way it
    is never a pass.
    """

    def __init__(self, reason: str) -> None:
        super().__init__(reason)
        self.reason = reason


# --------------------------------------------------------------------------- zone maths


def _is_imaginary(naive: datetime, zone: ZoneInfo) -> bool:
    """Did this local wall clock never happen? True inside a spring-forward gap.

    PEP 495's round trip: a wall time that does not survive local → UTC → local is
    one the zone skipped."""
    aware = naive.replace(tzinfo=zone)
    return aware.astimezone(UTC).astimezone(zone).replace(tzinfo=None) != naive


def first_real_instant(naive: datetime, zone: ZoneInfo) -> datetime:
    """The earliest instant a local wall clock names, as UTC.

    ``fold=0`` for an ambiguous time — the *earlier* of the two passes, which is the
    conservative reading of a deadline. For an imaginary time (inside a
    spring-forward gap) the deadline takes effect when the gap closes, so the search
    walks forward to the first wall clock the zone actually has. No real market puts
    a gate closure inside a transition, but a check about DST correctness does not
    get to assume that.
    """
    aware = naive.replace(tzinfo=zone)
    if not _is_imaginary(naive, zone):
        return aware.astimezone(UTC)
    for step in range(1, 24 * 60):
        probe = naive + timedelta(minutes=step)
        if not _is_imaginary(probe, zone):
            return probe.replace(tzinfo=zone).astimezone(UTC)
    return aware.astimezone(UTC)  # no real zone has a 24-hour gap


def day_start_utc(day: date, zone: ZoneInfo) -> datetime:
    """The instant a local delivery day begins."""
    return first_real_instant(datetime.combine(day, time(0, 0)), zone)


def periods_in_local_day(day: date, zone: ZoneInfo, minutes: int) -> int:
    """How many delivery periods that local day really has.

    Derived from the zone's own offsets, never from ``24 * 60 // minutes``: Europe/Oslo
    2024-10-27 is 25 hours (25 hourly / 100 quarter-hourly periods) and 2024-03-31 is
    23 (23 / 92)."""
    span = day_start_utc(day + timedelta(days=1), zone) - day_start_utc(day, zone)
    return int(span.total_seconds() // (minutes * 60))


def _day_hours(day: date, zone: ZoneInfo) -> float:
    span = day_start_utc(day + timedelta(days=1), zone) - day_start_utc(day, zone)
    return span.total_seconds() / 3600


def _transition_note(day: date, zone: ZoneInfo, minutes: int) -> str:
    """``""`` on an ordinary day; otherwise which transition made it short or long,
    naming the wall clock involved so the analyst can find the row."""
    hours = _day_hours(day, zone)
    if hours == 24:
        return ""
    naive = datetime.combine(day, time(0, 0))
    steps = int(24 * 60 / minutes) + 4
    if hours > 24:
        for step in range(steps):
            probe = naive + timedelta(minutes=step * minutes)
            if probe.date() == day and is_ambiguous_local(probe, zone):
                return f"DST fall-back: {probe:%H:%M} occurs twice, so the day is {hours:g} hours"
        return f"DST fall-back: the local day is {hours:g} hours"
    for step in range(steps):
        probe = naive + timedelta(minutes=step * minutes)
        if probe.date() == day and _is_imaginary(probe, zone):
            return (
                f"DST spring-forward: {probe:%H:%M} does not exist, so the day is {hours:g} hours"
            )
    return f"DST spring-forward: the local day is {hours:g} hours"


def _fmt(instant: datetime, zone: ZoneInfo) -> str:
    """An instant as local wall clock *with* its offset — the only honest way to
    print a time in evidence about time zones."""
    return instant.astimezone(zone).isoformat(timespec="seconds")


# --------------------------------------------------------------------------- headers


_BRACKETED = re.compile(r"[(\[{]\s*([^)\]}]+?)\s*[)\]}]\s*$")
_TOKENS = re.compile(r"[^0-9A-Za-z%/]+")


def _norm(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", text.lower()).strip()


def _known_unit(text: str) -> bool:
    candidate = text.strip().upper()
    return bool(candidate) and candidate in UNIT_TO_BASE


def split_header(header: str) -> tuple[str, str | None]:
    """``("volume", "mw")`` from ``volume_mw``; ``("Price", "EUR/MWh")`` from
    ``Price (EUR/MWh)``; ``("delivery_start", None)`` from ``delivery_start``.

    A bid file states its units in its headers, so the check has to read them — and
    having read them, addressing columns by name rather than by letter is free. The
    grammar is deliberately small and explicit: a trailing bracketed group, a
    trailing ``<currency>_<energy>`` pair, or a trailing single unit token.
    """
    text = header.strip()
    bracketed = _BRACKETED.search(text)
    if bracketed:
        inner = bracketed.group(1)
        if _known_unit(inner):
            return text[: bracketed.start()].strip(), inner
        return text, None
    tokens = [token for token in _TOKENS.split(text) if token]
    if not tokens:
        return text, None
    if len(tokens) >= 2:
        pair = f"{tokens[-2]}/{tokens[-1]}"
        if _known_unit(pair):
            return " ".join(tokens[:-2]), pair
    if len(tokens) >= 2 and _known_unit(tokens[-1]):
        return " ".join(tokens[:-1]), tokens[-1]
    return text, None


def _find_column(headers: list[str], name: str) -> int:
    """The single column whose header names ``name``, ignoring case, separators and
    any unit suffix. Zero matches or two are a stated refusal, never a guess."""
    target = _norm(name)
    hits = [
        index
        for index, header in enumerate(headers)
        if _norm(header) == target or _norm(split_header(header)[0]) == target
    ]
    if not hits:
        listed = ", ".join(repr(h) for h in headers[:20]) or "(none)"
        more = f" (+{len(headers) - 20} more)" if len(headers) > 20 else ""
        raise _CheckBlocked(
            f"no column named {name!r}; the artifact's headers are: {listed}{more}. "
            "Name the column as it appears in the file, or rename the header."
        )
    if len(hits) > 1:
        clashing = ", ".join(repr(headers[i]) for i in hits)
        raise _CheckBlocked(
            f"{len(hits)} columns match {name!r} ({clashing}); rename one so the "
            "reference is unambiguous."
        )
    return hits[0]


# --------------------------------------------------------------------------- reading


@dataclass
class _Table:
    """A bid/schedule artifact as headers plus ``(file row, values)`` pairs. The file
    row is carried because an evidence line that cannot say *which row* costs the
    analyst the search this check was supposed to save."""

    headers: list[str] = field(default_factory=list)
    rows: list[tuple[int, tuple[object, ...]]] = field(default_factory=list)


def _read_xlsx(path: Path, sheet: str | None, header_row: int) -> _Table:
    try:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise _CheckBlocked(f"workbook could not be opened: {exc}") from exc
    try:
        if sheet is not None:
            if sheet not in workbook.sheetnames:
                raise _CheckBlocked(
                    f"no sheet named {sheet!r}; the workbook has: "
                    f"{', '.join(repr(s) for s in workbook.sheetnames)}"
                )
            worksheet = workbook[sheet]
        else:
            worksheet = workbook.active
        table = _Table()
        for index, values in enumerate(worksheet.iter_rows(values_only=True), start=1):
            if index < header_row:
                continue
            if index == header_row:
                table.headers = ["" if value is None else str(value) for value in values]
                continue
            if all(value is None for value in values):
                continue
            table.rows.append((index, tuple(values)))
            if len(table.rows) > MAX_ROWS:
                raise _CheckBlocked(
                    f"artifact has more than {MAX_ROWS} data rows — that is over a leap "
                    "year of quarter-hourly periods. Split it by delivery period and re-run."
                )
        return table
    finally:
        workbook.close()


def _read_csv(path: Path, header_row: int) -> _Table:
    table = _Table()
    # utf-8-sig: Excel's "CSV UTF-8" export leads with a BOM, and a BOM glued to the
    # first header is a column the caller can never name.
    try:
        handle = path.open("r", encoding="utf-8-sig", newline="")
    except OSError as exc:
        raise _CheckBlocked(f"csv could not be opened: {exc}") from exc
    with handle:
        try:
            for index, values in enumerate(csv.reader(handle), start=1):
                if index < header_row:
                    continue
                if index == header_row:
                    table.headers = [value.strip() for value in values]
                    continue
                if not any(value.strip() for value in values):
                    continue
                table.rows.append((index, tuple(values)))
                if len(table.rows) > MAX_ROWS:
                    raise _CheckBlocked(
                        f"artifact has more than {MAX_ROWS} data rows — that is over a leap "
                        "year of quarter-hourly periods. Split it by delivery period and "
                        "re-run."
                    )
        except (csv.Error, UnicodeDecodeError) as exc:
            raise _CheckBlocked(f"csv could not be parsed: {exc}") from exc
    return table


def read_artifact(path: Path, sheet: str | None, header_row: int) -> _Table:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return _read_xlsx(path, sheet, header_row)
    if suffix == ".csv":
        return _read_csv(path, header_row)
    raise _CheckBlocked(
        f"unsupported artifact type {suffix or '(no extension)'!r}; this check reads "
        f"{' and '.join(SUPPORTED_SUFFIXES)}"
    )


# --------------------------------------------------------------------------- parsing


def _quote(value: object) -> str:
    text = "" if value is None else str(value).strip()
    return text[:MAX_QUOTE] + ("…" if len(text) > MAX_QUOTE else "")


def _cell(values: tuple[object, ...], index: int) -> object:
    return values[index] if index < len(values) else None


def parse_timestamp(raw: object) -> datetime | None:
    """A delivery-period start, or ``None`` when the cell is not one.

    A ``datetime`` straight out of openpyxl (naive), or an ISO-8601 string that may
    carry a UTC offset. Numbers are deliberately *not* accepted: an Excel serial in
    an unformatted cell is exactly the kind of value a lenient parser turns into the
    wrong day."""
    if isinstance(raw, datetime):
        return raw
    if isinstance(raw, date):
        return datetime.combine(raw, time(0, 0))
    if isinstance(raw, str):
        text = raw.strip()
        if not text:
            return None
        try:
            return datetime.fromisoformat(text)
        except ValueError:
            return None
    return None


def parse_fold(raw: object) -> int | None:
    """``0``/``1`` from a fold column, or ``None`` when it is neither."""
    if isinstance(raw, bool):
        return None
    if isinstance(raw, int):
        return raw if raw in (0, 1) else None
    if isinstance(raw, float):
        return int(raw) if raw in (0.0, 1.0) else None
    if isinstance(raw, str):
        text = raw.strip()
        return int(text) if text in ("0", "1") else None
    return None


# --------------------------------------------------------------------------- units


def _volume_finding(header: str, declared: str | None, minutes: int) -> UnitFinding:
    """Is the volume column's unit usable against a period of ``minutes``?

    The provable statement — provable from the header and the market alone, with no
    knowledge of the numbers — is the conversion factor. ``MW`` over a 60-minute MTU
    *is* ``MWh``; over 15 minutes it is four times it. A file that labels one and
    means the other is the classic silent bug, and naming the factor is what turns
    "looks fine" into a decision.
    """
    unit = declared if declared is not None else split_header(header)[1]
    finding = UnitFinding(column="volume", header=header, unit=unit, declared=declared is not None)
    if unit is None:
        finding.outcome = "warn"
        finding.detail = (
            f"The volume column {header!r} carries no unit and the spec declares none, so "
            "MW cannot be told from MWh. Over a "
            f"{minutes}-minute market time unit they differ by a factor of "
            f"{60 / minutes:g}. Put the unit in the header (volume_mwh) or set `volume_unit`."
        )
        return finding
    known = UNIT_TO_BASE.get(unit.strip().upper())
    if known is None:
        finding.outcome = "fail"
        finding.detail = (
            f"Unrecognised volume unit {unit!r}; use a power unit (MW/kW/GW) or an energy "
            "unit (MWh/kWh/GWh)."
        )
        return finding
    if known.dimension == "power":
        if minutes == 60:
            finding.detail = (
                f"The volume column is in {unit} (power); over a 60-minute market time unit "
                "that is numerically MWh, so the values need no conversion at this resolution."
            )
            return finding
        factor = 60 / minutes
        finding.outcome = "warn"
        finding.detail = (
            f"The volume column is in {unit} (power) but the market settles energy per "
            f"{minutes}-minute period: MWh = {unit} ÷ {factor:g}. Anything reading this file "
            f"as MWh will be {factor:g}x out. State the unit you mean."
        )
        return finding
    if known.dimension == "energy":
        converted, note = convert(1.0, unit, "MWh")
        finding.detail = (
            f"The volume column is in {unit} (energy per period), which is what the market "
            "settles." + ("" if note is None else f" Scale to MWh: x{converted:g}.")
        )
        return finding
    finding.outcome = "fail"
    finding.detail = (
        f"The volume column's unit {unit!r} is {known.dimension}; a delivery volume must be "
        "power (MW) or energy (MWh)."
    )
    return finding


def _price_finding(header: str, declared: str | None, market: MarketSpec) -> UnitFinding:
    """Is the price column denominated in what the market actually clears in?

    A different currency is a **refusal**, following the precedent
    ``services/reconciliation.py`` set: converting NOK/MWh into EUR/MWh needs a rate
    for a particular hour, which is an input the analyst owns and a gate never
    guesses. It is emphatically not an "unknown unit" — the unit is known, and that
    is why the answer can be this specific.
    """
    unit = declared if declared is not None else split_header(header)[1]
    finding = UnitFinding(column="price", header=header, unit=unit, declared=declared is not None)
    if unit is None:
        finding.outcome = "warn"
        finding.detail = (
            f"The price column {header!r} carries no unit and the spec declares none, so it "
            f"cannot be checked against {market.label}'s settlement currency "
            f"({market.currency}). Put it in the header (price_eur_mwh) or set `price_unit`."
        )
        return finding
    known = UNIT_TO_BASE.get(unit.strip().upper())
    if known is None:
        finding.outcome = "fail"
        finding.detail = (
            f"Unrecognised price unit {unit!r}; use <currency>/MWh or <currency>/kWh, e.g. "
            f"{market.currency}/MWh."
        )
        return finding
    if known.dimension != "price_energy":
        finding.outcome = "fail"
        finding.detail = (
            f"The price column's unit {unit!r} is {known.dimension}; a bid price is "
            f"<currency> per energy, e.g. {market.currency}/MWh."
        )
        return finding
    if known.currency != market.currency:
        finding.outcome = "fail"
        finding.detail = (
            f"The price column is denominated in {known.currency} but {market.label} clears "
            f"in {market.currency}. No FX rate will be invented — restate the prices in "
            f"{market.currency}/MWh, or select a market that clears in {known.currency}."
        )
        return finding
    converted, note = convert(1.0, unit, f"{market.currency}/MWh")
    finding.detail = f"The price column is in {unit}, the currency {market.label} clears in." + (
        "" if note is None else f" Scale to {market.currency}/MWh: x{converted:g}."
    )
    return finding


# --------------------------------------------------------------------------- the check


@dataclass
class _Resolved:
    """Every delivery period the artifact resolved to, plus what was learned on the
    way: whether the file disambiguated its own local times, and where it did not."""

    instants: dict[datetime, list[int]] = field(default_factory=dict)
    violations: list[PeriodViolation] = field(default_factory=list)
    naive_rows: int = 0
    aware_rows: int = 0
    fold_column_used: bool = False


def _record(out: _Resolved, instant: datetime, file_row: int, stamp: str) -> None:
    """Index one row's delivery period, or report a collision.

    The collision test is on the resolved **instant**, not on the text: two rows that
    say ``02:00+02:00`` and ``02:00+02:00`` and two that say ``02:00`` with folds 0
    and 0 are the same defect, and neither is visible by comparing strings."""
    existing = out.instants.setdefault(instant, [])
    if existing:
        out.violations.append(
            PeriodViolation(
                period=stamp,
                kind="duplicate",
                row=file_row,
                detail=(
                    f"row {file_row}: this bids the same delivery period as row "
                    f"{existing[0]}. One row per period."
                ),
            )
        )
        return
    existing.append(file_row)


class MarketRulesCheck:
    """The registered gate. ``id`` is the handle a
    :class:`~workbench_server.models.validation.ValidationSpec` names to run it."""

    id = "market_rules"

    async def run(self, ctx: ValidationContext) -> list[EvidenceItem]:
        try:
            spec = MarketCheckSpec.model_validate(ctx.params)
        except ValidationError as exc:
            return [_fail(f"invalid market-check spec: {exc.error_count()} error(s)")]
        market = MARKETS.get(spec.market)
        if market is None:
            return [
                _fail(
                    f"unknown market {spec.market!r}; the catalog holds "
                    f"{', '.join(market_ids())}. Market rules are server-owned and selected "
                    "by id — an artifact never states its own gate closure."
                )
            ]
        try:
            return self._run(spec, market, ctx)
        except _CheckBlocked as blocked:
            log.warning(
                "market_check.blocked",
                artifact=spec.artifact,
                market=spec.market,
                reason=blocked.reason,
            )
            return [_fail(f"{spec.artifact}: {blocked.reason}", market=market.id)]

    # ---- the body -----------------------------------------------------------

    def _run(
        self, spec: MarketCheckSpec, market: MarketSpec, ctx: ValidationContext
    ) -> list[EvidenceItem]:
        zone = ZoneInfo(market.timezone)
        path = self._resolve(ctx.root, spec.artifact)
        if path is None:
            raise _CheckBlocked("path escapes the workspace root")
        if not path.is_file():
            raise _CheckBlocked("artifact not found")

        table = read_artifact(path, spec.sheet, spec.header_row)
        if not table.headers:
            raise _CheckBlocked(f"no header row at row {spec.header_row}")
        if not table.rows:
            raise _CheckBlocked(f"no data rows below the header at row {spec.header_row}")

        ts_index = _find_column(table.headers, spec.timestamp_column)
        volume_index = _find_column(table.headers, spec.volume_column)
        price_index = (
            None if spec.price_column is None else _find_column(table.headers, spec.price_column)
        )
        fold_index = (
            None if spec.fold_column is None else _find_column(table.headers, spec.fold_column)
        )

        resolved = self._resolve_periods(spec, table, zone, ts_index, fold_index)
        if not resolved.instants:
            raise _CheckBlocked(
                f"none of the {len(table.rows)} data rows carried a usable delivery-period "
                f"start in column {table.headers[ts_index]!r} — nothing could be validated"
            )

        # `_coverage` returns one row per delivery day, oldest first, and `instants`
        # is non-empty by the guard above — so days[0] is the *binding* day: the one
        # whose gate closes first.
        days = self._coverage(spec, market, zone, resolved)
        gate = self._gate_closure(spec, market, zone, days[0].day, resolved)
        # The finest resolution any day in the artifact uses. A file spanning an MTU
        # switch must be judged by the strictest of the two: MW and MWh part company
        # at the shorter period, and that is the reading that can be wrong.
        finest = min(RESOLUTION_MINUTES[day.resolution] for day in days)
        units = [_volume_finding(table.headers[volume_index], spec.volume_unit, finest)]
        if price_index is not None:
            units.append(_price_finding(table.headers[price_index], spec.price_unit, market))

        report, truncation = self._report(
            spec, market, gate, days, units, resolved, len(table.rows)
        )
        ref = ctx.store_payload("artifact", report)

        evidence = [_gate_evidence(gate, market, spec)]
        evidence.append(_grid_evidence(spec, market, days, resolved, truncation, ref))
        evidence.extend(_day_evidence(spec, market, days))
        evidence.append(_ambiguity_evidence(spec, market, days, resolved))
        evidence.extend(
            EvidenceItem(
                kind="artifact",
                label=f"{finding.column} unit ({spec.artifact})",
                outcome=finding.outcome,
                detail=finding.detail,
            )
            for finding in units
        )
        log.info(
            "market_check.completed",
            artifact=spec.artifact,
            market=market.id,
            product=spec.product,
            days=len(days),
            violations=len(resolved.violations),
        )
        return evidence

    def _resolve(self, root: Path, relative: str) -> Path | None:
        """Jail the artifact path against the workspace root (the ``safe_path`` rule,
        inline: a check is handed the root, not the ``Workspace``)."""
        candidate = (root / relative).resolve()
        if candidate != root and root not in candidate.parents:
            return None
        return candidate

    # ---- rows → instants ----------------------------------------------------

    def _resolve_periods(
        self,
        spec: MarketCheckSpec,
        table: _Table,
        zone: ZoneInfo,
        ts_index: int,
        fold_index: int | None,
    ) -> _Resolved:
        """Turn each row's delivery-period start into a UTC instant, or into a named
        violation. Nothing is dropped silently."""
        out = _Resolved(fold_column_used=fold_index is not None)
        seen_local: dict[datetime, int] = {}
        for file_row, values in table.rows:
            raw = _cell(values, ts_index)
            parsed = parse_timestamp(raw)
            if parsed is None:
                out.violations.append(
                    PeriodViolation(
                        period=_quote(raw) or "(empty)",
                        kind="unparseable",
                        row=file_row,
                        detail=(
                            f"row {file_row}: {_quote(raw) or '(empty)'} is not an ISO-8601 "
                            "delivery-period start (e.g. 2024-10-27T02:00:00+02:00)."
                        ),
                    )
                )
                continue
            if parsed.tzinfo is not None:
                self._resolve_aware(out, parsed, zone, file_row)
                continue
            self._resolve_naive(out, spec, parsed, zone, file_row, seen_local, values, fold_index)
        return out

    def _resolve_aware(
        self, out: _Resolved, parsed: datetime, zone: ZoneInfo, file_row: int
    ) -> None:
        """A timestamp that carries its own UTC offset — the shape that needs no
        guessing. The offset is still checked against the zone: ``+03:00`` is not a
        thing Europe/Oslo has ever been, and a bid stamped with one is off by an hour
        in a way no downstream system will notice."""
        out.aware_rows += 1
        instant = parsed.astimezone(UTC)
        actual = instant.astimezone(zone).utcoffset()
        if actual != parsed.utcoffset():
            out.violations.append(
                PeriodViolation(
                    period=parsed.isoformat(timespec="seconds"),
                    kind="invalid_offset",
                    row=file_row,
                    detail=(
                        f"row {file_row}: offset {parsed.utcoffset()} is not the offset "
                        f"{zone.key} had at that instant ({actual}). The timestamp names a "
                        "different moment than it appears to."
                    ),
                )
            )
            return
        _record(out, instant, file_row, parsed.isoformat(timespec="seconds"))

    def _resolve_naive(
        self,
        out: _Resolved,
        spec: MarketCheckSpec,
        local: datetime,
        zone: ZoneInfo,
        file_row: int,
        seen_local: dict[datetime, int],
        values: tuple[object, ...],
        fold_index: int | None,
    ) -> None:
        """A bare local wall clock: the common shape, and the one that needs the
        zone's help. Order of appearance decides the fold **only where the zone
        genuinely repeats that hour** — the rule ``ReconciliationCheck`` already
        ships, for the same reason: row order alone cannot tell the second pass of a
        fall-back hour from a row someone pasted twice."""
        out.naive_rows += 1
        if _is_imaginary(local, zone):
            out.violations.append(
                PeriodViolation(
                    period=local.isoformat(timespec="seconds"),
                    kind="nonexistent",
                    row=file_row,
                    detail=(
                        f"row {file_row}: {local:%Y-%m-%dT%H:%M} does not exist in {zone.key} — "
                        "the clocks jump forward over it on that date. No delivery period "
                        "starts there."
                    ),
                )
            )
            return
        occurrence = seen_local.get(local, 0)
        seen_local[local] = occurrence + 1
        ambiguous = is_ambiguous_local(local, zone)
        stamp = local.isoformat(timespec="seconds")
        if fold_index is not None:
            fold = parse_fold(_cell(values, fold_index))
            if fold is None:
                out.violations.append(
                    PeriodViolation(
                        period=stamp,
                        kind="ambiguous",
                        row=file_row,
                        detail=(
                            f"row {file_row}: the fold column ({spec.fold_column}) must be 0 "
                            f"or 1, got {_quote(_cell(values, fold_index)) or '(empty)'}."
                        ),
                    )
                )
                return
            if fold == 1 and not ambiguous:
                out.violations.append(
                    PeriodViolation(
                        period=stamp,
                        kind="ambiguous",
                        row=file_row,
                        detail=(
                            f"row {file_row}: fold 1 was given for {local:%Y-%m-%dT%H:%M}, but "
                            f"that wall clock occurs once in {zone.key} — there is no second "
                            "occurrence."
                        ),
                    )
                )
                return
        elif occurrence == 0:
            fold = 0
        elif occurrence == 1 and ambiguous:
            fold = 1  # the genuine second pass of a fall-back hour
        else:
            # A repeat the zone does not justify — a third occurrence, or a second at
            # a wall clock that happens once. Reported rather than indexed: reading it
            # as a fold would bless corrupt input as a DST artefact, which is exactly
            # the mistake `ReconciliationCheck` refuses to make.
            out.violations.append(
                PeriodViolation(
                    period=stamp,
                    kind="duplicate",
                    row=file_row,
                    detail=(
                        f"row {file_row}: {local:%Y-%m-%dT%H:%M} already appeared "
                        f"{occurrence} time(s)"
                        + (
                            f", and {zone.key} repeats that wall clock exactly twice (the DST "
                            "fall-back hour)"
                            if ambiguous
                            else f", and it is not an ambiguous local time in {zone.key}"
                        )
                        + ". Remove the duplicate row, or disambiguate the rows with a UTC "
                        "offset or a fold column."
                    ),
                )
            )
            return
        _record(out, local.replace(tzinfo=zone, fold=fold).astimezone(UTC), file_row, stamp)

    # ---- per-day arithmetic -------------------------------------------------

    def _coverage(
        self, spec: MarketCheckSpec, market: MarketSpec, zone: ZoneInfo, resolved: _Resolved
    ) -> list[DayCoverage]:
        """Group the resolved instants by local delivery day and check each day's
        grid. Grouping by day is what makes the DST arithmetic possible: a period
        count is only meaningful against a *particular* local day."""
        by_day: dict[date, set[datetime]] = {}
        for instant in resolved.instants:
            by_day.setdefault(instant.astimezone(zone).date(), set()).add(instant)

        declared: date | None = None
        if spec.delivery_date is not None:
            try:
                declared = date.fromisoformat(spec.delivery_date)
            except ValueError as exc:
                raise _CheckBlocked(
                    f"delivery_date {spec.delivery_date!r} is not an ISO date (YYYY-MM-DD)"
                ) from exc

        out: list[DayCoverage] = []
        for day in sorted(by_day):
            row = market.resolution_for(day)
            if row is None:
                raise _CheckBlocked(
                    f"the catalog does not cover delivery date {day.isoformat()} for "
                    f"{market.id} — its rules start at "
                    f"{market.resolutions[0].effective_from.isoformat()}"
                )
            minutes = RESOLUTION_MINUTES[row.resolution]
            start = day_start_utc(day, zone)
            expected = periods_in_local_day(day, zone, minutes)
            grid = {start + timedelta(minutes=minutes * step) for step in range(expected)}
            observed = by_day[day]
            for instant in sorted(observed - grid):
                offset = (instant - start).total_seconds() / 60
                resolved.violations.append(
                    PeriodViolation(
                        period=_fmt(instant, zone),
                        kind="misaligned",
                        row=resolved.instants[instant][0],
                        detail=(
                            f"{_fmt(instant, zone)} is {offset:g} minutes into "
                            f"{day.isoformat()}, which is not a multiple of the "
                            f"{row.resolution} market time unit {market.label} settles on."
                        ),
                    )
                )
            for instant in sorted(grid - observed):
                resolved.violations.append(
                    PeriodViolation(
                        period=_fmt(instant, zone),
                        kind="missing",
                        detail=(
                            f"no row for delivery period {_fmt(instant, zone)} — the "
                            f"{day.isoformat()} grid needs all {expected} of them."
                        ),
                    )
                )
            if declared is not None and day != declared:
                resolved.violations.append(
                    PeriodViolation(
                        period=day.isoformat(),
                        kind="misaligned",
                        detail=(
                            f"the spec declares delivery day {declared.isoformat()} but the "
                            f"artifact also carries {len(observed)} period(s) for "
                            f"{day.isoformat()}."
                        ),
                    )
                )
            out.append(
                DayCoverage(
                    day=day.isoformat(),
                    resolution=row.resolution,
                    expected=expected,
                    observed=len(observed),
                    dst_transition=_transition_note(day, zone, minutes),
                )
            )
        return out

    # ---- the deadline -------------------------------------------------------

    def _gate_closure(
        self,
        spec: MarketCheckSpec,
        market: MarketSpec,
        zone: ZoneInfo,
        binding_day: str,
        resolved: _Resolved,
    ) -> GateClosureFinding:
        """The deadline verdict for the binding (earliest) delivery day.

        Day-ahead: one auction gate, ``day_ahead_gate_closure`` local wall-clock on
        D-1. Intraday: a rolling per-period deadline, so the binding one is the
        earliest delivery period's.
        """
        day = date.fromisoformat(binding_day)
        if spec.product == "intraday":
            earliest = min(resolved.instants)
            gate_utc = earliest - timedelta(minutes=market.intraday_lead_minutes)
        else:
            naive = datetime.combine(
                day - timedelta(days=market.day_ahead_gate_closure_days_before),
                market.day_ahead_gate_closure,
            )
            gate_utc = first_real_instant(naive, zone)
        finding = GateClosureFinding(
            market=market.id,
            product=spec.product,
            delivery_day=binding_day,
            gate_closure_local=_fmt(gate_utc, zone),
            gate_closure_utc=gate_utc.isoformat(timespec="seconds"),
        )
        if spec.submitted_at is None:
            return finding
        finding.submitted_input = spec.submitted_at
        try:
            parsed = datetime.fromisoformat(spec.submitted_at.strip())
        except ValueError:
            finding.outcome = "fail"
            return finding
        if parsed.tzinfo is not None:
            submitted = parsed.astimezone(UTC)
        else:
            # The documented reading: a bare wall clock is the analyst's own, in the
            # market's zone. The *other* reading is computed too — see below.
            finding.naive_input = True
            submitted = first_real_instant(parsed, zone)
            as_utc = parsed.replace(tzinfo=UTC)
            finding.readings_disagree = (submitted >= gate_utc) != (as_utc >= gate_utc)
        finding.submitted_utc = submitted.isoformat(timespec="seconds")
        if submitted >= gate_utc:
            finding.outcome = "fail"
        elif finding.readings_disagree:
            finding.outcome = "warn"
        else:
            finding.outcome = "pass"
        if spec.product == "intraday" and finding.outcome == "fail":
            for instant, rows in sorted(resolved.instants.items()):
                deadline = instant - timedelta(minutes=market.intraday_lead_minutes)
                if submitted >= deadline:
                    resolved.violations.append(
                        PeriodViolation(
                            period=_fmt(instant, zone),
                            kind="after_intraday_gate",
                            row=rows[0],
                            detail=(
                                f"delivery period {_fmt(instant, zone)} closed at "
                                f"{_fmt(deadline, zone)} "
                                f"({market.intraday_lead_minutes} min lead time)."
                            ),
                        )
                    )
        return finding

    # ---- the payload --------------------------------------------------------

    def _report(
        self,
        spec: MarketCheckSpec,
        market: MarketSpec,
        gate: GateClosureFinding,
        days: list[DayCoverage],
        units: list[UnitFinding],
        resolved: _Resolved,
        rows: int,
    ) -> tuple[MarketComplianceReport, EvidenceTruncation | None]:
        total = len(resolved.violations)
        window = resolved.violations
        truncated: EvidenceTruncation | None = None
        if total > MAX_VIOLATIONS:
            window = resolved.violations[:MAX_VIOLATIONS]
            truncated = EvidenceTruncation(
                shown=MAX_VIOLATIONS,
                total=total,
                detail=(
                    f"showing {MAX_VIOLATIONS} of {total} delivery-period violations; "
                    "validate one delivery day per run to see the rest"
                ),
            )
        report = MarketComplianceReport(
            artifact=spec.artifact,
            market=market.id,
            market_label=market.label,
            timezone=market.timezone,
            product=spec.product,
            gate_closure=gate,
            days=days,
            units=units,
            rows=rows,
            violations=window,
            truncated=truncated,
        )
        return report, truncated


# --------------------------------------------------------------------------- evidence


def _fail(reason: str, market: str | None = None) -> EvidenceItem:
    """A single ``fail`` line — never a silent skip, never a blank result a reader
    has to interpret (AXI shape 2)."""
    label = "market-rules compliance" + (f" ({market})" if market else "")
    return EvidenceItem(kind="artifact", label=label, outcome="fail", detail=reason)


def _gate_evidence(
    gate: GateClosureFinding, market: MarketSpec, spec: MarketCheckSpec
) -> EvidenceItem:
    """The headline. Kind ``gate`` because it is one — a deadline that either had
    passed or had not."""
    product = "intraday" if spec.product == "intraday" else "day-ahead"
    label = f"{product} gate closure ({market.id})"
    if gate.submitted_input is None:
        detail = (
            f"Not evaluated: the spec carries no `submitted_at`. {market.label} closes at "
            f"{gate.gate_closure_local} for delivery day {gate.delivery_day}. Supply the "
            "submission timestamp (ISO-8601, ideally with a UTC offset) to check it."
        )
        return EvidenceItem(kind="gate", label=label, outcome="skipped", detail=detail)
    if gate.submitted_utc is None:
        detail = (
            f"`submitted_at` {gate.submitted_input!r} is not an ISO-8601 timestamp, so the "
            f"{gate.gate_closure_local} deadline could not be checked. Use e.g. "
            "2024-10-26T11:59:59+02:00."
        )
        return EvidenceItem(kind="gate", label=label, outcome="fail", detail=detail)
    naive_note = (
        " The timestamp carries no UTC offset and was read as local wall-clock in "
        f"{market.timezone}."
        if gate.naive_input and not gate.readings_disagree
        else ""
    )
    if gate.outcome == "fail":
        detail = (
            f"Submitted {gate.submitted_input} — at or after {market.label}'s gate closure of "
            f"{gate.gate_closure_local} for delivery day {gate.delivery_day}. A bid at the "
            "closing instant is already late. Move the submission earlier, or bid the next "
            f"delivery day.{naive_note}"
        )
    elif gate.outcome == "warn":
        detail = (
            f"Submitted {gate.submitted_input} with no UTC offset, and the two readings "
            f"disagree: as local {market.timezone} it is before {market.label}'s "
            f"{gate.gate_closure_local} gate closure, as UTC it is not. A UTC-naive comparison "
            "against a local gate is the classic gate-closure bug — put an offset on "
            "`submitted_at` (…+02:00) or a trailing Z."
        )
    else:
        detail = (
            f"Submitted {gate.submitted_input}, before {market.label}'s gate closure of "
            f"{gate.gate_closure_local} for delivery day {gate.delivery_day}.{naive_note}"
        )
    return EvidenceItem(kind="gate", label=label, outcome=gate.outcome, detail=detail)


def _grid_evidence(
    spec: MarketCheckSpec,
    market: MarketSpec,
    days: list[DayCoverage],
    resolved: _Resolved,
    truncated: EvidenceTruncation | None,
    ref: str,
) -> EvidenceItem:
    """Alignment, missing, duplicate, nonexistent and bad-offset periods — one line,
    with the whole table behind ``payload_ref``."""
    total = len(resolved.violations)
    periods = len(resolved.instants)
    resolutions = ", ".join(sorted({day.resolution for day in days}))
    if total == 0:
        detail = (
            f"All {periods} delivery periods sit on {market.label}'s {resolutions} grid across "
            f"{len(days)} delivery day(s). No misaligned, missing or duplicated periods."
        )
        outcome: CheckOutcome = "pass"
    else:
        first = resolved.violations[0]
        detail = (
            f"{total} delivery-period violation(s) against {market.label}'s {resolutions} grid. "
            f"First: {first.detail}"
        )
        if truncated is not None:
            detail += f" {truncated.detail}."
        outcome = "fail"
    return EvidenceItem(
        kind="artifact",
        label=f"delivery-period grid ({spec.artifact})",
        outcome=outcome,
        detail=detail,
        payload_ref=ref,
    )


def _day_evidence(
    spec: MarketCheckSpec, market: MarketSpec, days: list[DayCoverage]
) -> list[EvidenceItem]:
    """The DST line. One ``fail`` per delivery day whose period count is wrong,
    bounded, plus a roll-up; one ``pass`` line when every day checks out."""
    bad = [day for day in days if day.observed != day.expected]
    if not bad:
        named = ", ".join(
            f"{day.day}: {day.expected}"
            + (f" ({day.dst_transition})" if day.dst_transition else "")
            for day in days[:MAX_NAMED_DAYS]
        )
        shapes = named + (
            f", +{len(days) - MAX_NAMED_DAYS} more" if len(days) > MAX_NAMED_DAYS else ""
        )
        return [
            EvidenceItem(
                kind="artifact",
                label=f"delivery-day period count ({spec.artifact})",
                outcome="pass",
                detail=(
                    f"Every delivery day carries the period count {market.timezone} gives it — "
                    f"{shapes}."
                ),
            )
        ]
    out = [
        EvidenceItem(
            kind="artifact",
            label=f"delivery-day period count ({day.day})",
            outcome="fail",
            detail=(
                f"{day.day} has {day.expected} {day.resolution} delivery periods in "
                f"{market.timezone}"
                + (f" — {day.dst_transition}" if day.dst_transition else "")
                + f"; the artifact carries {day.observed}. "
                + (
                    "A day around a DST transition is not 24 hours long, and a schedule that "
                    "assumes it is will be short or long by exactly the transition."
                    if day.dst_transition
                    else "Add the missing periods, or remove the extra ones."
                )
            ),
        )
        for day in bad[:MAX_DAY_LINES]
    ]
    withheld = len(bad) - MAX_DAY_LINES
    if withheld > 0:
        out.append(
            EvidenceItem(
                kind="artifact",
                label=f"delivery-day period count, {withheld} more ({spec.artifact})",
                outcome="fail",
                detail=(
                    f"{len(bad)} of {len(days)} delivery days have the wrong period count; the "
                    f"{MAX_DAY_LINES} earliest are named above and {withheld} more are not. "
                    "Validate one delivery day per run to see them."
                ),
            )
        )
    return out


def _ambiguity_evidence(
    spec: MarketCheckSpec,
    market: MarketSpec,
    days: list[DayCoverage],
    resolved: _Resolved,
) -> EvidenceItem:
    """Did the artifact say which of a repeated hour it meant?"""
    label = f"local-time disambiguation ({spec.artifact})"
    fallback = [day for day in days if day.dst_transition.startswith("DST fall-back")]
    if resolved.naive_rows == 0:
        detail = (
            f"All {resolved.aware_rows} delivery timestamps carry a UTC offset, so no local "
            f"time in {market.timezone} needed disambiguating."
        )
        return EvidenceItem(kind="artifact", label=label, outcome="pass", detail=detail)
    if not fallback:
        detail = (
            f"No delivery day in the artifact has a DST fall-back in {market.timezone}, so no "
            "local wall clock occurs twice and there is nothing to disambiguate."
        )
        return EvidenceItem(kind="artifact", label=label, outcome="pass", detail=detail)
    named = ", ".join(day.day for day in fallback[:MAX_NAMED_DAYS])
    if resolved.fold_column_used:
        detail = (
            f"{named} repeats a wall clock in {market.timezone}, and the artifact's "
            f"`{spec.fold_column}` column says which occurrence each row means."
        )
        return EvidenceItem(kind="artifact", label=label, outcome="pass", detail=detail)
    detail = (
        f"{named} repeats a wall clock in {market.timezone} ({fallback[0].dst_transition}), and "
        f"{resolved.naive_rows} delivery timestamp(s) carry no UTC offset and no fold column. "
        "They were read in order of appearance — first row = the earlier (summer-time) pass. "
        "Put an offset on the timestamps, or add a fold column, so the file says which hour it "
        "means rather than relying on row order."
    )
    return EvidenceItem(kind="artifact", label=label, outcome="warn", detail=detail)
