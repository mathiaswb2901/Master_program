"""The workbook↔code numeric reconciliation gate — M6's first *domain*
:class:`~workbench_server.services.validation.ValidationCheck`.

It reads the addressed cells through the :class:`WorkbookReader` seam, which has two
implementations. :class:`DiskWorkbookReader` opens the ``.xlsx`` **directly with
openpyxl**, deterministically, on a machine with no Office — the CI path.
:class:`LiveComWorkbookReader` reads the *running* Excel that has the workbook docked,
through the #92 COM bridge, on the shell backend's single apartment thread.

**Which one runs is not a preference; it is a gate.** A workbook docked in a panel with
an hour of unsaved edits in it makes the file on disk a different workbook — measured
against a real Excel: live ``B1`` 9999.0 while disk ``B1`` is still 1234.5, and a
formula cell Excel never calculated has *no* cached value on disk at all. A check that
read the file there would report **pass** about numbers it did not read, which is the
silent green the whole milestone exists to refuse. So before a single cell is opened the
check asks the office host whether the workbook is docked and, if it is, whether it is
saved:

===========  ==================  ==================  =====================================
Docked?      ``Workbook.Saved``  Live read?          Verdict
===========  ==================  ==================  =====================================
no           —                   —                   read disk, sourced ``file @mtime``
yes          ``True``            either              read live if available, else disk
yes          ``False``           yes                 read **live**, sourced ``live @…``
yes          ``False``           **no**              **blocked**, naming the remedy
===========  ==================  ==================  =====================================

The fourth row is the whole point, and it is why ``CheckOutcome`` grew a ``blocked``
member: refusing *with a sentence* is worth something, and refusing by returning no
evidence at all (the only way to reach ``blocked`` before) throws that sentence away.

Three domain failure modes are first-class here, not comments:

* **Units are compared, not assumed** (:func:`convert`). A value read in kWh against an
  MWh expectation, or MWh against MW, cannot pass silently — the x1000 that hides inside
  a spreadsheet diff is caught, and a legitimate unit change is *named* in the evidence.
  Prices are currency-aware: NOK/SEK/DKK are first-class alongside EUR, scaling within a
  currency is a named multiply, and comparing *across* currencies is an explicit refusal
  (:class:`CrossCurrency`) rather than an invented FX rate.
* **Time-indexed rows are aligned by local wall-clock time** (:func:`_align_time_rows`),
  telling the repeated 02:00 of a fall-back DST day apart from an ordinary row that was
  pasted twice by asking the *zone* whether that wall clock is genuinely ambiguous
  (:func:`_is_ambiguous`) — never by order of appearance alone. A naive positional join
  that drops or invents the duplicated hour is refused, and so is a duplicate row
  masquerading as a fold. A timestamp that carries a **UTC offset** is refused rather
  than normalised (:class:`OffsetNotAllowed`): the two occurrences of a fall-back hour
  differ *only* by their offset, so stripping it hands them the same lookup key and the
  fold gate falls over on the one day it exists for.
* **No look-ahead in the pairing.** Rows are matched by their own timestamp value, never
  by position, so a missing or extra row surfaces as an unmatched expectation rather than
  a comparison against the wrong hour.

The check returns **one grouped** :class:`EvidenceItem` (kind ``numeric``) whose outcome
is the worst across all comparisons, and stores the whole
:class:`~workbench_server.models.reconciliation.ReconciliationReport` as a bounded payload
via :meth:`ValidationContext.store_payload` — the table never rides the result or the bus.
"""

from __future__ import annotations

from collections.abc import Iterable, Iterator, Sequence
from datetime import UTC, datetime
from itertools import islice
from pathlib import Path
from typing import NamedTuple, Protocol
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

import openpyxl
import structlog
from openpyxl.utils import column_index_from_string
from pydantic import ValidationError

from workbench_server.models.office_bridge import LiveWorkbookStatus
from workbench_server.models.reconciliation import (
    CellComparison,
    ExpectedValue,
    ReadSource,
    ReconciliationReport,
    ReconciliationSpec,
    TimeExpectation,
    TimeIndexSpec,
    Tolerance,
)
from workbench_server.models.validation import (
    CheckOutcome,
    EvidenceItem,
    EvidenceTruncation,
)

# A *value* helper, not the office host. `naive_local` is the one implementation
# of "drop the offset COM attaches, keep the wall clock", and the live reader
# applies it again at its own seam so the invariant — nothing tz-aware ever
# crosses `WorkbookReader` — holds for any bridge, not only the COM one. Two
# copies of that rule would be two places for it to drift.
from workbench_server.services.office_host.office_com import naive_local
from workbench_server.services.validation import ValidationContext

log = structlog.get_logger()

#: Cap on the comparison rows carried inside one stored report. The table is a
#: bounded window, worst-first; beyond this, :class:`EvidenceTruncation` says how
#: much was cut and how to narrow the run (AXI shape 1).
MAX_COMPARISONS = 200

#: Cap on the *named* duplicate-timestamp evidence lines. One paste accident can
#: repeat thousands of rows, and a fail per row would bury the grouped verdict and
#: blow the result's token budget; beyond this the rest are rolled into one line
#: that says how many were withheld and how to see them (AXI shape 1).
MAX_DUPLICATE_LINES = 20

#: Character clip on the offending timestamp quoted back in the offset finding. The
#: cell content is the workbook's, so it is bounded before it reaches the evidence.
_OFFSET_SAMPLE_CLIP = 40

#: Hard ceiling on the rows **one column read pulls**, live or from disk. A year
#: of quarter-hours is 35,040 rows, so this clears the realistic worst case with
#: room; a taller column is cut there. Truncation can only ever *lose* a row,
#: which surfaces as an unmatched expectation ("alignment gap") — it can never
#: manufacture agreement, which is the property that makes the cap safe.
#:
#: It bounds *both* readers, and that is the point rather than tidiness. It was
#: the live read's cap alone (``MAX_LIVE_ROWS``) while the disk read had none, so
#: a 60,000-row column reconciled 50,000 rows docked and 60,000 undocked — the
#: same defect :func:`column_data_rows` exists to prevent, one level up. A cap
#: that bounds the *read* is also the only kind that helps: slicing the report
#: afterwards costs the same read.
MAX_COLUMN_ROWS = 50_000

#: Least-to-most severe, for rolling per-cell outcomes up into one grouped
#: verdict. ``blocked`` sorts above ``fail`` for the same reason the risk ramp
#: does: a check that could not judge is worse than one that judged and
#: disagreed.
_OUTCOME_SEVERITY: dict[CheckOutcome, int] = {
    "pass": 0,
    "skipped": 1,
    "warn": 2,
    "fail": 3,
    "blocked": 4,
}


class _Unit(NamedTuple):
    """One known unit: what it measures, how it scales to that dimension's base
    unit, and — for monetary units only — which currency it is denominated in."""

    dimension: str
    factor: float
    currency: str | None = None


def _currency_units(code: str) -> dict[str, _Unit]:
    """The three monetary units of one currency: the bare amount and the two
    per-energy denominations an analyst actually types."""
    return {
        f"{code}/MWH": _Unit("price_energy", 1.0, code),
        f"{code}/KWH": _Unit("price_energy", 1e3, code),
        code: _Unit("money", 1.0, code),
    }


#: unit (upper-cased) → (dimension, factor to the dimension's base unit, currency).
#: Two units are comparable iff they share a dimension **and**, when monetary, the
#: same currency; the factor turns a value in the unit into the base unit (MWh for
#: energy, MW for power, <ccy>/MWh for price, <ccy> for money). A x1000 conversion
#: (kWh↔MWh, NOK/kWh↔NOK/MWh) therefore cannot happen by accident — it is a named,
#: explicit multiply or it is a fail.
#:
#: The Nordic currencies are first-class, not an afterthought: Nord Pool publishes
#: NO/SE/DK zone prices in both EUR and the local currency, and a model denominated
#: in NOK/MWh is the normal case for this product's users — it must reconcile with a
#: named x1000 like the EUR path does, not fall off the table as an "unknown unit".
#: Deliberately hand-rolled rather than pulled from a units library: the whole value
#: here is the *refusal* rules (dimension, currency), which a general converter would
#: happily paper over.
_UNIT_TO_BASE: dict[str, _Unit] = {
    "MWH": _Unit("energy", 1.0),
    "KWH": _Unit("energy", 1e-3),
    "GWH": _Unit("energy", 1e3),
    "WH": _Unit("energy", 1e-6),
    "MW": _Unit("power", 1.0),
    "KW": _Unit("power", 1e-3),
    "GW": _Unit("power", 1e3),
    "W": _Unit("power", 1e-6),
    "%": _Unit("ratio", 1.0),
    "": _Unit("dimensionless", 1.0),
    **_currency_units("EUR"),
    **_currency_units("NOK"),
    **_currency_units("SEK"),
    **_currency_units("DKK"),
}


class UnitMismatch(Exception):
    """Raised when two units name different physical dimensions (energy vs power),
    or when an unknown unit cannot be matched by string equality."""


class CrossCurrency(UnitMismatch):
    """Raised when two *known, same-dimension* monetary units are denominated in
    different currencies.

    A subclass of :class:`UnitMismatch` so every existing caller keeps failing
    safe, but a distinct type because it is a different verdict with a different
    fix: nothing is unknown here and nothing is malformed — the gate is refusing
    to invent an FX rate. Converting NOK/MWh to EUR/MWh needs a rate *for a
    particular hour*, which is an input the analyst owns and the gate never
    guesses. Silently applying one would be the same class of hidden multiply the
    unit table exists to catch, only worse: wrong by a number nobody chose.
    """


def convert(value: float, from_unit: str, to_unit: str) -> tuple[float, str | None]:
    """Convert ``value`` from ``from_unit`` into ``to_unit``.

    Returns ``(converted, note)`` where ``note`` names the conversion when one
    actually happened (so a x1000 is always visible in the evidence) and is ``None``
    for an identity. Raises :class:`UnitMismatch` across incompatible dimensions or
    for an unknown unit that does not string-match the target, and the narrower
    :class:`CrossCurrency` for two prices in different currencies.
    """
    a = from_unit.strip().upper()
    b = to_unit.strip().upper()
    if a == b:
        return value, None
    known_a = _UNIT_TO_BASE.get(a)
    known_b = _UNIT_TO_BASE.get(b)
    if known_a is None or known_b is None:
        raise UnitMismatch(f"unknown unit in {from_unit!r} vs {to_unit!r}")
    if known_a.dimension != known_b.dimension:
        raise UnitMismatch(f"{from_unit} is {known_a.dimension}, {to_unit} is {known_b.dimension}")
    if known_a.currency != known_b.currency:
        raise CrossCurrency(
            f"{from_unit} is in {known_a.currency} and {to_unit} is in {known_b.currency}; "
            "no FX rate will be invented — restate the expectation in "
            f"{known_a.currency}, or convert it in your own code at the rate you meant"
        )
    converted = value * known_a.factor / known_b.factor
    return converted, f"converted {value:g} {from_unit} → {converted:g} {to_unit}"


def within_tolerance(expected: float, actual: float, tol: Tolerance) -> bool:
    """A match iff the absolute delta is within ``abs`` *or* within ``rel*|expected|``.

    Never a naive ``==``: at least one band is guaranteed set by
    :class:`Tolerance`'s own validator."""
    delta = abs(actual - expected)
    if tol.abs is not None and delta <= tol.abs:
        return True
    return tol.rel is not None and delta <= tol.rel * abs(expected)


def column_data_rows(rows: Iterable[tuple[object, object]]) -> list[tuple[object, object]]:
    """Where a time-index column's data ends: every row before the first one whose
    **both** cells are empty.

    **The single statement of that rule**, and it is called by both
    implementations of :class:`WorkbookReader` — the disk reader over a lazy row
    generator (so it still stops reading rather than scanning a sheet it would
    throw away) and the live reader over the window a bridge handed back. One
    function rather than one rule per reader, because two copies is exactly what
    drifted: the live path used to keep the whole used range and trim only the
    *tail*, so a workbook with a blank row 3 and more hours at row 4 reconciled
    three rows live and one row from disk. Same file, same spec, two comparison
    sets — decided by whether the workbook happened to be docked when the check
    ran. That hollows out the provenance this module exists to print: naming
    ``live`` or ``file`` is only worth something when the two read the same rows.

    The rule kept is the disk reader's, unchanged, and the choice is not
    arbitrary. Cutting at the gap can only ever *lose* rows, and a lost row
    surfaces as an unmatched expectation ("alignment gap") — the same property
    that makes :data:`MAX_COLUMN_ROWS` safe. It can never manufacture agreement,
    which is the one thing a validation gate must not do.
    """
    out: list[tuple[object, object]] = []
    for pair in rows:
        if pair == (None, None):
            break
        out.append(pair)
    return out


class WorkbookReader(Protocol):
    """The seam the disk reader and the live COM reader both satisfy.

    A cell value is a number when numeric, a ``datetime`` when the cell holds a
    timestamp, ``None`` when empty, or a ``str`` when the cell holds text (which the
    gate treats as unreadable-for-numbers).

    **Every implementation says where it read.** :meth:`provenance` is not
    bookkeeping — with two readers in the tree, a comparison table that cannot name
    its source is a table nobody can act on, and the report carries the answer so a
    reader of the *evidence* never has to ask.

    **And no implementation hands a tz-aware value across.** The disk reader yields
    naive local wall clock because that is what openpyxl stores; the live reader
    normalises at its own boundary. That is what keeps
    :class:`OffsetNotAllowed` protecting the case it was written for — an
    offset-bearing *string* in the workbook — instead of firing on every row of a
    live read.

    **And both end a column in the same place.** :meth:`column_pairs` returns the
    rows :func:`column_data_rows` keeps, never a per-reader notion of where the
    data stops — a reader that decided that for itself would make the same
    workbook reconcile differently docked and undocked.
    """

    def cell_value(self, sheet: str | None, cell: str) -> object: ...

    def column_pairs(
        self, sheet: str | None, ts_column: str, value_column: str, start_row: int
    ) -> list[tuple[object, object]]: ...

    def provenance(self) -> ReadSource: ...

    def close(self) -> None: ...


class DiskWorkbookReader:
    """Reads cached, computed values from an ``.xlsx`` with no Office installed.

    ``data_only=True`` reads the values Excel last cached — the correct source for
    reconciling *numbers* rather than formula strings. ``read_only=True`` keeps a large
    workbook cheap **as long as it is read the way a read-only sheet wants to be read**:
    streamed, forwards, once (:meth:`column_pairs`). It has no random access, so the
    convenient-looking ``ws["A7"]`` is the expensive call, not the cheap one.
    A workbook that was never opened in Excel has no cached values; the
    check turns the resulting all-``None`` into a **fail that names the fix** rather
    than a silent green, and :attr:`ReadSource.cached_values` records it so the
    evidence can tell "a workbook of zeros" from "a workbook nobody has opened".

    Formerly ``OpenpyxlReader``; renamed when the second implementation arrived, so
    the two read like the pair they are rather than one named after its library and
    one after its transport.
    """

    def __init__(self, path: Path, read_at: datetime) -> None:
        self._wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        self._read_at = read_at
        # Naive local, like every other timestamp in this module.
        self._mtime = datetime.fromtimestamp(path.stat().st_mtime)
        self._saw_value = False

    def close(self) -> None:
        self._wb.close()

    def provenance(self) -> ReadSource:
        return ReadSource(
            kind="file",
            read_at=self._read_at,
            mtime=self._mtime,
            cached_values=self._saw_value,
        )

    def _sheet(self, sheet: str | None) -> object:
        if sheet is None:
            return self._wb.active
        if sheet not in self._wb.sheetnames:
            raise KeyError(sheet)
        return self._wb[sheet]

    def cell_value(self, sheet: str | None, cell: str) -> object:
        ws = self._sheet(sheet)
        value = ws[cell].value  # type: ignore[index]
        if value is not None:
            self._saw_value = True
        if value is None or isinstance(value, int | float | str | datetime):
            return value
        return str(value)

    def column_pairs(
        self, sheet: str | None, ts_column: str, value_column: str, start_row: int
    ) -> list[tuple[object, object]]:
        """The two columns from ``start_row`` down, in **one streamed forward pass**.

        This is the only way to read a read-only worksheet that is not quadratic,
        and the reason is worth writing down because the shape that replaced it
        looks so innocent. ``ReadOnlyWorksheet`` has no random access: every
        ``ws["A7"]`` re-opens the sheet's zip member and re-runs the SAX parse
        **from row 1** to find that one cell. The previous loop asked for two
        addresses per data row, so a column of N rows cost 2N parses of an
        O(N)-row document — measured in this repo's venv at 500 rows 4.3 s,
        1,000 rows 17.5 s, 2,000 rows 55.1 s, four-fold per doubling. An annual
        hourly series is 8,760 rows (twenty-odd minutes) and a multi-year
        quarter-hourly one is 35,040 (hours), inside a gate that is supposed to
        run on save. It also made the class docstring's "``read_only=True`` keeps
        a large workbook cheap" false. After: 0.15 s at 10,000 rows, 0.48 s at
        35,040.

        ``iter_rows`` is the documented streamed read: one SAX pass, rows handed
        out in order, and — measured against a real ``.xlsx`` — missing rows are
        padded with ``None`` cells, so a gap in the data still arrives as the
        ``(None, None)`` that :func:`column_data_rows` stops on and a column that
        starts below ``start_row`` still stops immediately. Same rows in, same
        rows out; the pass is lazy, so a sheet whose data ends at row 40 is still
        not scanned to row 1,048,576.

        The band is ``min(ts, value)`` to ``max(ts, value)`` — two columns that
        are far apart cost the columns between them per row, which is bounded and
        linear, where a second pass would be another whole parse. The read is
        bounded by :data:`MAX_COLUMN_ROWS`, the same ceiling the live reader gets,
        and it bounds the *read* rather than the report.
        """
        ws = self._sheet(sheet)
        # A column letter that is not one (``"A1"``, ``""``) raises here, and
        # `_compare_time_rows` turns that into a named fail per expectation. The
        # address form it replaced silently read a *different* cell instead.
        ts_index = column_index_from_string(ts_column)
        value_index = column_index_from_string(value_column)
        first, last = min(ts_index, value_index), max(ts_index, value_index)

        def rows() -> Iterator[tuple[object, object]]:
            banded = ws.iter_rows(  # type: ignore[attr-defined]
                min_row=start_row, min_col=first, max_col=last, values_only=True
            )
            for row in islice(banded, MAX_COLUMN_ROWS):
                yield row[ts_index - first], row[value_index - first]

        pairs = column_data_rows(rows())
        if pairs:
            self._saw_value = True
        return pairs


class LiveWorkbookHost(Protocol):
    """The narrow view of the Office host the front gate needs, and nothing more.

    Declared here and satisfied by
    :class:`~workbench_server.services.office_host.service.OfficeHostService`, so
    this module never imports the host — the same posture ``services/gates.py``
    takes with its ``SlotLocator``. It also means the whole four-row table is
    drivable in a unit test from a twenty-line stub.
    """

    def docked_workbook(self, path: str) -> str | None: ...

    async def workbook_status(self, path: str) -> LiveWorkbookStatus: ...

    async def read_workbook_cells(
        self, path: str, sheet: str | None, cells: Sequence[str]
    ) -> list[object]: ...

    async def read_workbook_columns(
        self,
        path: str,
        sheet: str | None,
        ts_column: str,
        value_column: str,
        start_row: int,
        max_rows: int,
    ) -> list[tuple[object, object]]: ...


class LiveComWorkbookReader:
    """Reads the *running* Excel that has the workbook docked, over the COM bridge.

    **A snapshot, not a live cursor.** Everything the spec addresses is fetched in
    one pass before any comparison runs, and the synchronous
    :class:`WorkbookReader` methods then serve from that snapshot. Three reasons,
    all of them load-bearing: COM calls are ``await``ed on the shell backend's one
    apartment thread and the protocol is sync; a table compared cell-by-cell against
    a workbook the user is *still typing into* would be internally inconsistent, row
    3 read a second after row 2; and one round trip per run is what keeps the gate
    from being slower than the edit it is checking.

    **Normalised at this seam.** Every snapshot value goes through
    :func:`~workbench_server.services.office_host.office_com.naive_local`, so a
    tz-aware datetime — which real COM returns, with an offset that is *not* this
    machine's zone — becomes the naive local wall clock the rest of the module is
    written in, and never reaches ``_as_local_datetime``'s offset refusal.

    **And cut at this seam.** A bridge returns the window it was asked for and
    decides nothing about it; :func:`column_data_rows` — the disk reader's own
    rule, called here — says where the column's data ends. Putting it here rather
    than in each bridge is what makes it structurally impossible for a *future*
    bridge to reconcile a different set of rows than the file would.
    """

    def __init__(
        self,
        cells: dict[tuple[str | None, str], object],
        columns: dict[tuple[str | None, str, str, int], list[tuple[object, object]]],
        source: ReadSource,
    ) -> None:
        self._cells = cells
        self._columns = columns
        self._source = source

    @classmethod
    async def load(
        cls,
        host: LiveWorkbookHost,
        host_path: str,
        spec: ReconciliationSpec,
        status: LiveWorkbookStatus,
        read_at: datetime,
    ) -> LiveComWorkbookReader:
        """Take the snapshot. Anything the bridge refuses propagates — the caller
        decides whether that is a fall back to disk or a refusal, and on a dirty
        workbook it is always the latter."""
        addresses: dict[str | None, list[str]] = {}
        for exp in spec.expectations:
            sheet, addr = _split_cell(exp.cell)
            addresses.setdefault(sheet, []).append(addr)
        cells: dict[tuple[str | None, str], object] = {}
        for sheet, wanted in addresses.items():
            values = await host.read_workbook_cells(host_path, sheet, wanted)
            for addr, value in zip(wanted, values, strict=False):
                cells[(sheet, addr)] = naive_local(value)

        columns: dict[tuple[str | None, str, str, int], list[tuple[object, object]]] = {}
        ti = spec.time_index
        if ti is not None and ti.expectations:
            pairs = await host.read_workbook_columns(
                host_path,
                ti.sheet,
                ti.timestamp_column,
                ti.value_column,
                ti.start_row,
                MAX_COLUMN_ROWS,
            )
            # The bridge hands back the window it was asked for, verbatim; where
            # the *data* ends is this seam's decision and the disk reader's rule.
            columns[(ti.sheet, ti.timestamp_column, ti.value_column, ti.start_row)] = (
                column_data_rows((naive_local(ts), naive_local(value)) for ts, value in pairs)
            )
        return cls(
            cells,
            columns,
            ReadSource(
                kind="live",
                read_at=read_at,
                calculation=status.calculation,
                saved=status.saved,
            ),
        )

    def close(self) -> None:
        """Nothing to release: the instance belongs to the host, and the snapshot
        is a dict. Here so the two readers are interchangeable at the call site."""

    def provenance(self) -> ReadSource:
        return self._source

    def cell_value(self, sheet: str | None, cell: str) -> object:
        key = (sheet, cell)
        if key not in self._cells:
            # The snapshot is taken from the spec, so a miss means the spec and the
            # read disagree about what was addressed. Raised rather than returned as
            # an empty cell, because "empty" is a verdict and this is a bug.
            raise KeyError(f"{cell} was not in the live snapshot")
        return self._cells[key]

    def column_pairs(
        self, sheet: str | None, ts_column: str, value_column: str, start_row: int
    ) -> list[tuple[object, object]]:
        key = (sheet, ts_column, value_column, start_row)
        if key not in self._columns:
            raise KeyError(f"{ts_column}/{value_column} was not in the live snapshot")
        return self._columns[key]


def _split_cell(cell: str) -> tuple[str | None, str]:
    """``Sheet1!D14`` → ``("Sheet1", "D14")``; ``D14`` → ``(None, "D14")``."""
    if "!" in cell:
        sheet, addr = cell.split("!", 1)
        return sheet or None, addr
    return None, cell


def _as_number(value: object) -> float | None:
    """A cell's value as a float, or ``None`` when it is empty or non-numeric.

    Booleans are *not* numbers here — a ``TRUE`` cell reconciled as ``1.0`` is a lie."""
    if isinstance(value, bool):
        return None
    if isinstance(value, int | float):
        return float(value)
    return None


def _tolerance_for(spec: ReconciliationSpec, cell: str) -> Tolerance:
    return spec.per_cell_tolerance.get(cell, spec.default_tolerance)


def _compare_one(
    address: str,
    label: str | None,
    expected: float,
    expected_unit: str,
    cell_unit: str,
    raw: object,
    tol: Tolerance,
) -> CellComparison:
    """Turn one expected/actual pair into a verdict, unit-aware."""
    actual = _as_number(raw)
    if actual is None:
        reason = "empty cell" if raw is None else f"non-numeric cell ({raw!r})"
        return CellComparison(
            cell=address,
            label=label,
            expected=expected,
            actual=None,
            unit=expected_unit,
            delta=None,
            outcome="fail",
            reason=reason,
        )
    try:
        converted, note = convert(actual, cell_unit, expected_unit)
    except CrossCurrency as exc:
        # Named apart from a unit mismatch on purpose: the units are both known and
        # both prices, so "unit mismatch" would send the reader looking for a typo.
        return CellComparison(
            cell=address,
            label=label,
            expected=expected,
            actual=actual,
            unit=expected_unit,
            delta=None,
            outcome="fail",
            reason=f"cross-currency comparison refused: {exc}",
        )
    except UnitMismatch as exc:
        return CellComparison(
            cell=address,
            label=label,
            expected=expected,
            actual=actual,
            unit=expected_unit,
            delta=None,
            outcome="fail",
            reason=f"unit mismatch: {exc}",
        )
    delta = converted - expected
    if within_tolerance(expected, converted, tol):
        return CellComparison(
            cell=address,
            label=label,
            expected=expected,
            actual=converted,
            unit=expected_unit,
            delta=delta,
            outcome="pass",
            reason=note,
        )
    band = f"outside tolerance (Δ {delta:g})"
    reason = f"{note}; {band}" if note else band
    return CellComparison(
        cell=address,
        label=label,
        expected=expected,
        actual=converted,
        unit=expected_unit,
        delta=delta,
        outcome="fail",
        reason=reason,
    )


def _time_address(exp: TimeExpectation) -> str:
    """The wall-clock label a time-row comparison is keyed by in the evidence."""
    return exp.timestamp + (f" (fold {exp.fold})" if exp.fold else "")


class OffsetNotAllowed(ValueError):
    """Raised when a timestamp that must be a naive local wall clock carries a UTC
    offset.

    A subclass of :class:`ValueError` so every existing caller keeps failing safe, but
    a distinct type because it is a different verdict with a different fix: nothing is
    malformed here — the value parses perfectly. The gate is refusing to *reinterpret*
    it, exactly as :class:`CrossCurrency` refuses to invent an FX rate.

    Stripping the offset is not a harmless normalisation on a fall-back day. The two
    02:00s of ``2024-10-27`` in ``Europe/Oslo`` differ **only** by their offset
    (``+02:00`` then ``+01:00``) — that is the idiomatic, DST-safe way a pandas or
    Python export writes them. Dropping it collapses both to the same naive wall clock
    with ``fold`` defaulting to 0, so the two carry the *same* lookup key: the genuine
    fold-1 row is never consulted and the second expectation is silently scored against
    the first one's value. The fold gate is defeated on precisely the day it exists to
    protect, and the run comes back green.
    """


#: The refusal an offset-bearing timestamp earns, on both sides of the comparison.
#: Kept short enough to survive the ``office_reconcile`` reason clip (140 chars) with
#: the lesson intact — a truncated lesson is not a lesson — and it never repeats the
#: offending value, which the comparison row already carries as its address.
_OFFSET_REFUSAL = (
    "timestamp carries a UTC offset; this field is naive local wall clock — "
    "drop the offset and use fold (0/1) to pick the fall-back occurrence"
)


def _parse_local(ts: str) -> datetime:
    """A naive local ISO timestamp string.

    Raises :class:`OffsetNotAllowed` when the string carries a UTC offset (or a ``Z``),
    and a plain ``ValueError`` when it is not an ISO timestamp at all. Both are
    ``ValueError``, which the callers already turn into a *named* per-row fail rather
    than an exception that sinks the run."""
    parsed = datetime.fromisoformat(ts)
    if parsed.tzinfo is not None:
        raise OffsetNotAllowed(_OFFSET_REFUSAL)
    return parsed


def _as_local_datetime(value: object) -> datetime | None:
    """One workbook cell as a naive local wall clock, or ``None`` when it is not a
    timestamp at all.

    Raises :class:`OffsetNotAllowed` for an offset-bearing value — a tz-aware
    ``datetime`` object or a string like ``2024-10-27T02:00:00+01:00`` (what
    ``df.index.astype(str)`` writes). Refused on this side too, and for the same
    reason: with the offset stripped, which of a fall-back day's two 02:00 rows is
    fold 0 would be decided by **row order alone**, and order of appearance is the one
    thing :func:`_align_time_rows` promises never to infer a fold from. A descending
    export would silently swap the two hours."""
    if isinstance(value, datetime):
        if value.tzinfo is not None:
            raise OffsetNotAllowed(_OFFSET_REFUSAL)
        return value
    if isinstance(value, str):
        try:
            return _parse_local(value)
        except OffsetNotAllowed:
            raise
        except ValueError:
            return None
    return None


def _is_ambiguous(local: datetime, zone: ZoneInfo) -> bool:
    """Is this naive local wall-clock time one the zone genuinely writes **twice**?

    True only on a fall-back transition — 2024-10-27 02:30 in ``Europe/Oslo`` is both
    the CEST 02:30 and the CET 02:30 an hour later. PEP 495's own test: the two folds
    of the same wall time carry different UTC offsets. The *imaginary* times of a
    spring-forward gap also disagree on offset, so they are excluded first by the
    round-trip — a wall time that does not survive local → UTC → local never happened,
    and a row carrying it is not a second occurrence of anything.
    """
    fold0 = local.replace(tzinfo=zone, fold=0)
    if fold0.astimezone(UTC).astimezone(zone).replace(tzinfo=None) != local:
        return False  # imaginary: inside the spring-forward gap
    return fold0.utcoffset() != local.replace(tzinfo=zone, fold=1).utcoffset()


class _AlignedRows(NamedTuple):
    """The wall-clock index, plus the timestamps that repeated when the zone says
    they should not have, plus the rows refused for carrying a UTC offset."""

    #: ``(local wall-clock, fold)`` → the value that row carried.
    values: dict[tuple[datetime, int], object]
    #: local wall-clock → how many rows carried it, for timestamps that repeated
    #: **without** being an ambiguous (fall-back) local time in the zone.
    duplicates: dict[datetime, int]
    #: How many rows were refused for carrying a UTC offset (see
    #: :class:`OffsetNotAllowed`), and the first one of them, for the evidence. A
    #: count and one sample rather than a list: a whole tz-aware column is 8,760
    #: rows and one finding covers them all.
    offset_rows: int = 0
    offset_sample: str | None = None


def _align_time_rows(pairs: list[tuple[object, object]], zone: ZoneInfo) -> _AlignedRows:
    """Map ``(local wall-clock, fold)`` → value, keyed by the timestamp the row
    *carries* and never by its position.

    A second row at the same wall-clock is a ``fold=1`` **only where ``zone`` actually
    repeats that hour**. That test is the whole point: order of appearance alone cannot
    tell the CET 02:00 of a fall-back day apart from an ordinary row someone pasted
    twice, and a gate that guesses will *match* a fold-1 expectation against corrupt
    input and report it reconciled. So a repeat at a non-ambiguous local time (and any
    third-and-later repeat, which no zone produces) is not indexed at all: it is
    reported to the caller as a duplicate row, and the expectation it would have
    satisfied gets no value to satisfy it.

    A row whose timestamp carries a UTC offset is **refused, not stripped**, and the
    count comes back to the caller as its own finding: normalising it away would put
    the fold decision back on row order, which is the guess this function exists to
    refuse.

    Aligning by value (not index) is what keeps a missing or extra row from silently
    shifting every later hour."""
    seen: dict[datetime, int] = {}
    values: dict[tuple[datetime, int], object] = {}
    duplicates: dict[datetime, int] = {}
    offset_rows = 0
    offset_sample: str | None = None
    for ts_raw, value in pairs:
        try:
            local = _as_local_datetime(ts_raw)
        except OffsetNotAllowed:
            offset_rows += 1
            if offset_sample is None:
                offset_sample = str(ts_raw)[:_OFFSET_SAMPLE_CLIP]
            continue
        if local is None:
            continue
        occurrence = seen.get(local, 0)
        seen[local] = occurrence + 1
        if occurrence == 0:
            values[(local, 0)] = value
            continue
        if occurrence == 1 and _is_ambiguous(local, zone):
            values[(local, 1)] = value  # the genuine second pass of a fall-back hour
            continue
        duplicates[local] = seen[local]
    return _AlignedRows(
        values=values,
        duplicates=duplicates,
        offset_rows=offset_rows,
        offset_sample=offset_sample,
    )


class ReconciliationCheck:
    """The registered gate. ``id`` is the handle a
    :class:`~workbench_server.models.validation.ValidationSpec` names to run it.

    ``host`` is the live-workbook seam. ``None`` — a server with no Office host,
    which is every CI runner — means the gate reads disk, exactly as it always
    did; nothing about the no-Office path changes.
    """

    id = "reconciliation"

    def __init__(self, host: LiveWorkbookHost | None = None) -> None:
        self._host = host

    async def run(self, ctx: ValidationContext) -> list[EvidenceItem]:
        try:
            spec = ReconciliationSpec.model_validate(ctx.params)
        except ValidationError as exc:
            return [self._spec_error(f"invalid reconciliation spec: {exc.error_count()} error(s)")]

        workbook = self._resolve(ctx.root, spec.workbook)
        if workbook is None:
            return [self._spec_error(f"workbook path escapes the workspace: {spec.workbook!r}")]
        if not workbook.is_file():
            return [self._spec_error(f"workbook not found: {spec.workbook}")]

        if not spec.expectations and not (spec.time_index and spec.time_index.expectations):
            return [self._spec_error("spec has no expectations to reconcile")]

        # Naive local wall clock, minted here — the module's contract for every
        # timestamp, and deliberately not read off a COM object.
        read_at = datetime.now().replace(microsecond=0)
        try:
            reader, gate_evidence = await self._acquire(spec, workbook, read_at)
        except _LiveReadBlocked as refusal:
            return [self._blocked(spec, refusal.reason)]
        except _CheckBlocked as unopenable:
            return [self._spec_error(unopenable.reason)]

        try:
            comparisons = self._compare_cells(spec, reader)
            time_rows, structural = self._compare_time_rows(spec, reader)
            comparisons += time_rows
        except _CheckBlocked as blocked:
            return [self._spec_error(blocked.reason)]
        finally:
            reader.close()

        # The grouped comparison line first (it is what a reader looks at), then any
        # structural findings about the rows themselves. Both feed `derive_risk`, so a
        # duplicated-row fail cannot be outvoted by comparisons that happened to match.
        return [self._grouped(spec, comparisons, ctx, reader), *gate_evidence, *structural]

    # ---- the front gate -----------------------------------------------------

    async def _acquire(
        self, spec: ReconciliationSpec, workbook: Path, read_at: datetime
    ) -> tuple[WorkbookReader, list[EvidenceItem]]:
        """Decide *what to read* before reading anything, and hand back the reader.

        The order here is the gate. The host is consulted first, always: once the
        ``.xlsx`` has been opened the decision has already been made, and a later
        edit that reordered these two lines would turn the whole thing back into a
        read-then-check that reports about the file. Raises
        :class:`_LiveReadBlocked` for the one case with no honest answer — docked,
        dirty, and no live read.
        """
        host_path = None if self._host is None else self._host.docked_workbook(spec.workbook)
        if host_path is None or self._host is None:
            # Not docked: the file *is* the workbook, and disk is the whole truth.
            return self._disk(spec, workbook, read_at), []

        try:
            status = await self._host.workbook_status(host_path)
        except Exception as exc:
            # Docked, and we could not find out whether it is clean. There is no
            # conservative disk fallback here: the question this failed to answer
            # is precisely "would reading the file be wrong?".
            log.warning("reconciliation.status_unreadable", workbook=spec.workbook, error=str(exc))
            raise _LiveReadBlocked(
                f"{spec.workbook} is open in Workbench, but its live state could not be read "
                f"({exc}). Workbench cannot tell whether it has unsaved changes, and "
                "reconciling the file on disk would judge numbers you may already have "
                "changed. Close it in Excel, or re-run from the desktop shell."
            ) from exc

        gate_evidence: list[EvidenceItem] = []
        if status.calculation != "done":
            # A third state, and not a pass: the formula cells hold numbers Excel
            # has not finished revising. The comparisons still run — they are
            # informative — but this line keeps the result off green.
            gate_evidence.append(
                EvidenceItem(
                    kind="numeric",
                    label=f"live workbook still calculating ({spec.workbook})",
                    outcome="skipped",
                    detail=(
                        f"Excel has not finished calculating {spec.workbook} (state: "
                        f"{status.calculation}), so its formula results are still moving. "
                        "Re-run in a moment for a verdict these numbers can carry."
                    ),
                )
            )

        try:
            reader = await LiveComWorkbookReader.load(self._host, host_path, spec, status, read_at)
        except Exception as exc:
            if not status.saved:
                log.warning(
                    "reconciliation.live_read_unavailable",
                    workbook=spec.workbook,
                    error=str(exc),
                )
                raise _LiveReadBlocked(
                    f"{spec.workbook} is open in Excel with unsaved changes and the live "
                    f"reader is not available here ({exc}). Save the workbook, or run this "
                    "from the desktop shell, and re-run — reconciling the file on disk "
                    "would judge numbers you have already changed."
                ) from exc
            # Clean and docked: the file and the instance agree, so disk is a
            # correct answer rather than a convenient one. Named as `file` in the
            # evidence either way.
            log.info(
                "reconciliation.live_read_declined",
                workbook=spec.workbook,
                error=str(exc),
            )
            return self._disk(spec, workbook, read_at), gate_evidence
        return reader, gate_evidence

    def _disk(
        self, spec: ReconciliationSpec, workbook: Path, read_at: datetime
    ) -> DiskWorkbookReader:
        try:
            return DiskWorkbookReader(workbook, read_at)
        except Exception as exc:  # a corrupt/locked workbook is a fail that names why
            log.warning("reconciliation.unreadable", workbook=spec.workbook, error=str(exc))
            raise _CheckBlocked(f"workbook could not be opened: {exc}") from exc

    def _resolve(self, root: Path, relative: str) -> Path | None:
        """Jail the workbook path against the workspace root (the ``safe_path``
        rule, inline: a check is handed the root, not the ``Workspace``)."""
        candidate = (root / relative).resolve()
        if candidate != root and root not in candidate.parents:
            return None
        return candidate

    def _compare_cells(
        self, spec: ReconciliationSpec, reader: WorkbookReader
    ) -> list[CellComparison]:
        out: list[CellComparison] = []
        empties = 0
        for exp in spec.expectations:
            sheet, addr = _split_cell(exp.cell)
            try:
                raw = reader.cell_value(sheet, addr)
            except Exception as exc:
                # The reader seam: a malformed address (e.g. a bare row "14", which
                # openpyxl resolves to a *tuple* of cells whose `.value` raises
                # AttributeError) or an unknown sheet is one isolated fail row that
                # names the reason — never an exception that sinks the whole run.
                log.warning("reconciliation.cell_unreadable", cell=exp.cell, error=str(exc))
                out.append(self._unreadable_cell(exp, f"cannot read {exp.cell}: {exc}"))
                empties += 1
                continue
            if raw is None:
                empties += 1
            out.append(
                _compare_one(
                    exp.cell,
                    exp.label,
                    exp.expected,
                    exp.unit,
                    exp.cell_unit if exp.cell_unit is not None else exp.unit,
                    raw,
                    _tolerance_for(spec, exp.cell),
                )
            )
        # A workbook never opened in Excel caches no values: every addressed cell
        # reads empty. Report that as its own blocked-style failure that names the
        # fix, rather than a wall of identical "empty cell" rows.
        if spec.expectations and empties == len(spec.expectations):
            raise _CheckBlocked(
                f"every addressed cell in {spec.workbook} is empty — open and save the "
                "workbook in Excel once so values are cached, then re-run"
            )
        return out

    def _compare_time_rows(
        self, spec: ReconciliationSpec, reader: WorkbookReader
    ) -> tuple[list[CellComparison], list[EvidenceItem]]:
        """The time-indexed comparisons, plus any structural evidence about the rows
        themselves (duplicated timestamps) that is not a per-expectation verdict."""
        ti = spec.time_index
        if ti is None or not ti.expectations:
            return [], []
        if spec.timezone is None:
            raise _CheckBlocked("time_index is set but `timezone` is missing")
        try:
            zone = ZoneInfo(spec.timezone)
        except (ZoneInfoNotFoundError, ValueError) as exc:
            raise _CheckBlocked(f"unknown timezone {spec.timezone!r}: {exc}") from exc

        try:
            pairs = reader.column_pairs(
                ti.sheet, ti.timestamp_column, ti.value_column, ti.start_row
            )
        except Exception as exc:
            # A bad sheet (KeyError) or a malformed column letter is isolated to the
            # time-index rows — each becomes a fail that names why — so the direct-cell
            # comparisons in the same run still land in the stored report.
            log.warning(
                "reconciliation.time_index_unreadable",
                workbook=spec.workbook,
                error=str(exc),
            )
            reason = (
                f"cannot read time-index columns ({ti.timestamp_column}/{ti.value_column}): {exc}"
            )
            return [
                CellComparison(
                    cell=_time_address(exp),
                    label=exp.label,
                    expected=exp.expected,
                    actual=None,
                    unit=exp.unit,
                    outcome="fail",
                    reason=reason,
                )
                for exp in ti.expectations
            ], []
        aligned = _align_time_rows(pairs, zone)
        structural = [
            *self._offset_evidence(spec, ti, aligned),
            *self._duplicate_evidence(spec, ti, aligned, zone),
        ]

        out: list[CellComparison] = []
        for exp in ti.expectations:
            address = _time_address(exp)
            try:
                local = _parse_local(exp.timestamp)
            except OffsetNotAllowed as exc:
                # Named apart from an unparseable string on purpose: this one parses
                # fine. Stripping the offset is what used to make the two occurrences
                # of a fall-back hour share a lookup key, so the fold-1 expectation
                # was scored against the fold-0 row and the run came back green.
                log.warning(
                    "reconciliation.offset_bearing_expectation",
                    workbook=spec.workbook,
                    # Not `timestamp=`: structlog's own TimeStamper writes that key
                    # last, so a field by that name never reaches the log line.
                    expectation=exp.timestamp,
                    timezone=spec.timezone,
                )
                out.append(
                    CellComparison(
                        cell=address,
                        label=exp.label,
                        expected=exp.expected,
                        actual=None,
                        unit=exp.unit,
                        outcome="fail",
                        reason=str(exc),
                    )
                )
                continue
            except ValueError:
                out.append(
                    CellComparison(
                        cell=address,
                        label=exp.label,
                        expected=exp.expected,
                        actual=None,
                        unit=exp.unit,
                        outcome="fail",
                        reason=f"unparseable timestamp {exp.timestamp!r}",
                    )
                )
                continue
            if exp.fold and not _is_ambiguous(local, zone):
                # A fold only exists where the zone repeats an hour. Asking for one
                # anywhere else is a spec error, and saying "alignment gap" here would
                # send the analyst hunting for a missing workbook row that was never
                # supposed to exist.
                out.append(
                    CellComparison(
                        cell=address,
                        label=exp.label,
                        expected=exp.expected,
                        actual=None,
                        unit=exp.unit,
                        outcome="fail",
                        reason=(
                            f"fold {exp.fold} was requested but {exp.timestamp} is not an "
                            f"ambiguous local time in {spec.timezone} — that wall clock "
                            "occurs once, so there is no second occurrence to match"
                        ),
                    )
                )
                continue
            key = (local, exp.fold)
            if key not in aligned.values:
                out.append(
                    CellComparison(
                        cell=address,
                        label=exp.label,
                        expected=exp.expected,
                        actual=None,
                        unit=exp.unit,
                        outcome="fail",
                        reason="no workbook row at this local time (alignment gap)",
                    )
                )
                continue
            cell_unit = ti.value_unit if ti.value_unit is not None else exp.unit
            out.append(
                _compare_one(
                    address,
                    exp.label,
                    exp.expected,
                    exp.unit,
                    cell_unit,
                    aligned.values[key],
                    _tolerance_for(spec, exp.timestamp),
                )
            )
        return out, structural

    def _offset_evidence(
        self, spec: ReconciliationSpec, ti: TimeIndexSpec, aligned: _AlignedRows
    ) -> list[EvidenceItem]:
        """One ``fail`` line for a timestamp column that carries UTC offsets.

        Its own :class:`EvidenceItem`, like the duplicate finding, because it is a
        statement about the *workbook* rather than about any one expectation — and
        because the alternative is worse than a fail: the refused rows are not indexed,
        so without this line the expectations at those hours would come back as a bare
        "alignment gap" and send the analyst hunting for rows that are sitting right
        there. One line for the whole column, not one per row: a tz-aware export is
        8,760 of them.
        """
        if aligned.offset_rows == 0:
            return []
        log.warning(
            "reconciliation.offset_bearing_column",
            workbook=spec.workbook,
            column=ti.timestamp_column,
            rows=aligned.offset_rows,
            sample=aligned.offset_sample,
            timezone=spec.timezone,
        )
        return [
            EvidenceItem(
                kind="numeric",
                label=f"timestamp column carries UTC offsets ({spec.workbook})",
                outcome="fail",
                detail=(
                    f"{aligned.offset_rows} row(s) in column {ti.timestamp_column} carry a "
                    f"UTC offset (first: {aligned.offset_sample!r}). The offset is not "
                    "stripped: dropping it would leave row order to decide which of the two "
                    f"02:00 rows of a {spec.timezone} fall-back day is the first pass, and "
                    "order of appearance is exactly what a fold must never be inferred from. "
                    "Those rows were not indexed, so expectations at those hours report an "
                    "alignment gap. Write the column as naive local wall clock in "
                    f"{spec.timezone} (the repeated hour twice, in order), then re-run."
                ),
            )
        ]

    def _duplicate_evidence(
        self,
        spec: ReconciliationSpec,
        ti: TimeIndexSpec,
        aligned: _AlignedRows,
        zone: ZoneInfo,
    ) -> list[EvidenceItem]:
        """One ``fail`` line per duplicated timestamp — the defect that used to be a
        silent ``fold=1``.

        It is its own :class:`EvidenceItem` rather than a comparison row because it is
        a statement about the *workbook*, not about any one expectation: the duplicate
        is a fail whether or not an expectation happens to address that hour, and a
        result that carries it can never derive ``pass``.

        Bounded at :data:`MAX_DUPLICATE_LINES` named lines: one paste accident can
        repeat thousands of rows, and the roll-up line that follows states how many
        were withheld rather than leaving the reader to wonder.
        """
        out: list[EvidenceItem] = []
        ordered = sorted(aligned.duplicates.items())
        for local, count in ordered[:MAX_DUPLICATE_LINES]:
            stamp = local.isoformat()
            log.warning(
                "reconciliation.duplicate_timestamp",
                workbook=spec.workbook,
                # `local_time`, not `timestamp`: structlog's TimeStamper owns that
                # key, so the duplicated hour this line exists to name was being
                # overwritten by the log's own clock and never appeared.
                local_time=stamp,
                count=count,
                timezone=spec.timezone,
            )
            if _is_ambiguous(local, zone):
                # A genuine fall-back hour, but repeated *more* than the two times the
                # zone writes it. Saying "not ambiguous" here would be a false statement
                # in the evidence, and would send the analyst to the wrong fix.
                why = (
                    f"{spec.timezone} repeats that wall clock exactly twice (the DST "
                    f"fall-back hour), so {count - 2} of these row(s) are duplicates"
                )
            else:
                why = (
                    f"it is not an ambiguous local time in {spec.timezone} — that wall "
                    "clock occurs once, so these are duplicate rows, not a DST fall-back "
                    "repeat"
                )
            out.append(
                EvidenceItem(
                    kind="numeric",
                    label=f"duplicate timestamp row ({spec.workbook})",
                    outcome="fail",
                    detail=(
                        f"{stamp} appears {count} times in column {ti.timestamp_column}, "
                        f"but {why}. The duplicates were not matched against any "
                        "expectation. Remove them (or fix the timestamps), then re-run."
                    ),
                )
            )
        withheld = len(ordered) - MAX_DUPLICATE_LINES
        if withheld > 0:
            out.append(
                EvidenceItem(
                    kind="numeric",
                    label=f"duplicate timestamp rows, {withheld} more ({spec.workbook})",
                    outcome="fail",
                    detail=(
                        f"{len(ordered)} distinct timestamps are duplicated in column "
                        f"{ti.timestamp_column}; the {MAX_DUPLICATE_LINES} earliest are "
                        f"named above and {withheld} more are not. De-duplicate the "
                        f"timestamp column in {spec.workbook} and re-run to see whether "
                        "any remain."
                    ),
                )
            )
        return out

    def _grouped(
        self,
        spec: ReconciliationSpec,
        comparisons: list[CellComparison],
        ctx: ValidationContext,
        reader: WorkbookReader,
    ) -> EvidenceItem:
        """One grouped ``numeric`` evidence line; the full table goes to the payload
        store and is named by ``payload_ref``.

        ``reader`` is here for one field: the report has to say which of the two
        it came from, and asking the reader that produced the values is the only
        way that answer cannot drift from the values."""
        total = len(comparisons)
        matched = sum(1 for c in comparisons if c.outcome == "pass")
        mismatched = sum(1 for c in comparisons if c.outcome == "fail")
        worst = max(
            (c.outcome for c in comparisons),
            key=lambda o: _OUTCOME_SEVERITY[o],
            default="pass",
        )

        # Worst-first, then bound the stored window (AXI shape 1).
        ordered = sorted(comparisons, key=lambda c: -_OUTCOME_SEVERITY[c.outcome])
        truncated: EvidenceTruncation | None = None
        window = ordered
        if total > MAX_COMPARISONS:
            window = ordered[:MAX_COMPARISONS]
            truncated = EvidenceTruncation(
                shown=MAX_COMPARISONS,
                total=total,
                detail=(
                    f"showing {MAX_COMPARISONS} of {total} comparisons, worst first; "
                    "reconcile fewer cells per run to see the rest"
                ),
            )
        source = reader.provenance()
        report = ReconciliationReport(
            workbook=spec.workbook,
            matched=matched,
            mismatched=mismatched,
            total=total,
            source=source,
            comparisons=window,
            truncated=truncated,
        )
        ref = ctx.store_payload("numeric", report)

        if mismatched == 0:
            detail = f"All {total} cells reconcile within tolerance."
        else:
            first_bad = next(c for c in ordered if c.outcome == "fail")
            detail = (
                f"{mismatched} of {total} cells mismatch beyond tolerance. "
                f"First: {first_bad.cell} in {spec.workbook}."
            )
        # The provenance rides the headline, not only the payload: a reader who
        # has to expand a table to find out which of two workbooks a green badge
        # is about does not have proof, they have a colour.
        return EvidenceItem(
            kind="numeric",
            label=f"workbook↔code reconciliation ({spec.workbook})",
            outcome=worst,
            detail=f"{detail} {describe_source(source)}",
            payload_ref=ref,
        )

    def _unreadable_cell(self, exp: ExpectedValue, reason: str) -> CellComparison:
        return CellComparison(
            cell=exp.cell,
            label=exp.label,
            expected=exp.expected,
            actual=None,
            unit=exp.unit,
            outcome="fail",
            reason=reason,
        )

    def _spec_error(self, reason: str) -> EvidenceItem:
        """A single ``fail`` evidence line — never a silent skip and never a
        blank result a reader has to interpret (AXI shape 2)."""
        return EvidenceItem(
            kind="numeric",
            label="workbook↔code reconciliation",
            outcome="fail",
            detail=reason,
        )

    def _blocked(self, spec: ReconciliationSpec, reason: str) -> EvidenceItem:
        """The front gate's refusal: **no numbers were read, and none will be**.

        Distinct from ``_spec_error``'s ``fail`` because it is a different verdict
        with a different fix. A ``fail`` says the workbook and the code disagree;
        this says nobody can honestly say whether they do, and names the two
        things that would change that. It carries no ``payload_ref`` — there is no
        comparison table, and inventing an empty one would suggest a read happened.
        """
        log.warning("reconciliation.blocked", workbook=spec.workbook, detail=reason)
        return EvidenceItem(
            kind="numeric",
            label=f"workbook↔code reconciliation blocked ({spec.workbook})",
            outcome="blocked",
            detail=reason,
        )


def describe_source(source: ReadSource) -> str:
    """One sentence naming where a report's numbers came from.

    Written for the *evidence line*, which is the one place a reader always sees
    — "read live at 14:02:11, calculation done, workbook unsaved" or "read from
    the file on disk, modified 13:41:08". The panel renders the same fields from
    the payload; this is the version that survives a collapsed expander.
    """
    if source.kind == "live":
        clean = "workbook saved" if source.saved else "workbook unsaved"
        return (
            f"Read live from the docked workbook at {source.read_at:%H:%M:%S}, "
            f"calculation {source.calculation}, {clean}."
        )
    when = "unknown time" if source.mtime is None else f"{source.mtime:%Y-%m-%d %H:%M:%S}"
    empty = "" if source.cached_values is not False else " — the file cached no values"
    return f"Read from the file on disk, modified {when}{empty}."


class _CheckBlocked(Exception):
    """Internal: a condition that stops the whole reconciliation with one reason.

    Reported as a ``fail``: the workbook *was* reached and could not be made to
    yield the numbers the spec named. Its sibling below is the other thing.
    """

    def __init__(self, reason: str) -> None:
        super().__init__(reason)
        self.reason = reason


class _LiveReadBlocked(Exception):
    """Internal: the front gate refused before a single cell was read.

    A different verdict from :class:`_CheckBlocked` and rendered as ``blocked``
    rather than ``fail`` — nothing here disagrees with anything. Workbench is
    holding a workbook open whose live numbers it cannot read, and the file on
    disk is known not to be those numbers.
    """

    def __init__(self, reason: str) -> None:
        super().__init__(reason)
        self.reason = reason
