"""The ``office_reconcile`` agent tool — the reconciliation gate made agent-usable.

An agent that just wrote numbers into a workbook checks them against its own computed
expectations and gets an honest pass/warn/fail with evidence. These drive the tool the
way a session hits it: through :func:`handle_office_reconcile`, against a *real*
``ValidationService`` with the real ``ReconciliationCheck`` registered and a real tiny
``.xlsx`` on disk — no fakes in the path under test. What is pinned here is the domain
behaviour (a kWh-vs-MWh 1000x is caught, an unreadable workbook is an honest fail, not a
crash), the AXI three shapes (a truncated worst-list states what it withheld and names
the full table), and the tool's own token budget (description and serialized result).
"""

import json
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl

from workbench_server.models.agent_reconcile import ReconcileSummary
from workbench_server.services.agent_tools import (
    MAX_DESCRIPTION_CHARS,
    OFFICE_RECONCILE,
    RECONCILE_WORST_N,
    allowed_tool_names,
    handle_office_reconcile,
)
from workbench_server.services.event_bus import EventBus
from workbench_server.services.reconciliation import ReconciliationCheck
from workbench_server.services.validation import ValidationService


def make_workbook(path: Path, cells: dict[str, Any], sheet: str = "Sheet1") -> None:
    """A tiny workbook with the given A1 cells populated (mirrors test_reconciliation)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet
    for addr, value in cells.items():
        ws[addr] = value
    wb.save(path)


def runner_for(root: Path) -> ValidationService:
    """A real ValidationService with the reconciliation check registered — the
    production wiring, minus the app."""
    service = ValidationService(root, EventBus())
    service.register(ReconciliationCheck())
    return service


def result_text(result: dict[str, Any]) -> str:
    text: str = result["content"][0]["text"]
    return text


def summary_of(result: dict[str, Any]) -> ReconcileSummary:
    return ReconcileSummary.model_validate_json(result_text(result))


# --------------------------------------------------------------------------- behaviour


async def test_a_matching_workbook_reconciles_clean(tmp_path: Path) -> None:
    """The happy path: every cell agrees within tolerance, risk pass, and the
    summary says all N reconciled — no mismatch to name."""
    make_workbook(tmp_path / "book.xlsx", {"B2": 12.5, "B3": 42.0, "B4": 7.25})
    runner = runner_for(tmp_path)
    args = {
        "workbook": "book.xlsx",
        "default_tolerance": {"abs": 0.01},
        "expectations": [
            {"cell": "B2", "expected": 12.5, "unit": "MWh"},
            {"cell": "B3", "expected": 42.0, "unit": "MWh"},
            {"cell": "B4", "expected": 7.25, "unit": "MWh"},
        ],
    }
    summary = summary_of(await handle_office_reconcile(runner, args))
    assert summary.risk == "pass"
    assert summary.passed == 3
    assert summary.failed == 0
    assert summary.total == 3
    assert summary.worst == []
    assert summary.withheld == 0
    assert "All 3 cells" in summary.summary
    # AXI shape 3: the clean answer still names where the full table lives.
    assert summary.validation_id in summary.next_step


async def test_a_kwh_vs_mwh_1000x_is_caught_and_named(tmp_path: Path) -> None:
    """The moat, exercised: the agent computed 12.5 MWh, but the workbook cell holds
    12500 — a raw kWh figure that landed in an MWh column, the single most common
    silent bug in this domain. Compared as MWh (as the column claims) it is 1000x
    high, so it cannot pass. Risk high, the worst mismatch named with its delta, and a
    next-step line that points at the fix. (When the cell is *honestly* declared kWh
    via cell_unit, the gate converts and it passes — that is the sibling case; here the
    column lies about its unit, which is exactly what must fail.)"""
    make_workbook(tmp_path / "book.xlsx", {"B2": 12500.0})
    runner = runner_for(tmp_path)
    args = {
        "workbook": "book.xlsx",
        "default_tolerance": {"abs": 0.01, "rel": 0.001},
        "expectations": [
            {"cell": "B2", "expected": 12.5, "unit": "MWh"},
        ],
    }
    summary = summary_of(await handle_office_reconcile(runner, args))
    assert summary.risk == "high"
    assert summary.failed == 1
    assert summary.total == 1
    assert len(summary.worst) == 1
    worst = summary.worst[0]
    assert worst.cell == "B2"
    assert worst.expected == 12.5
    assert worst.actual == 12500.0
    assert worst.delta is not None and worst.delta > 12000
    # The reason names the tolerance band — the evidence a human reads.
    assert worst.reason
    # AXI shape 3: end with the obvious next step and the named full table.
    assert "re-run" in summary.next_step
    assert summary.validation_id in summary.next_step


async def test_an_unreadable_workbook_is_an_honest_fail_not_a_crash(tmp_path: Path) -> None:
    """A workbook that is not there cannot be read: the tool must answer with an
    honest fail that names the reason, never a silent green and never an exception."""
    runner = runner_for(tmp_path)
    args = {
        "workbook": "missing.xlsx",
        "default_tolerance": {"abs": 0.01},
        "expectations": [{"cell": "B2", "expected": 1.0, "unit": "MWh"}],
    }
    summary = summary_of(await handle_office_reconcile(runner, args))
    # Not a pass — a validation that could not judge is high/blocked, never green.
    assert summary.risk in ("high", "blocked")
    assert summary.total == 0
    assert summary.worst == []
    # The reason is surfaced, not swallowed: the agent is told *why* it failed.
    assert "not found" in summary.summary.lower() or "workbook" in summary.summary.lower()


async def test_an_empty_workbook_cell_fails_rather_than_passes(tmp_path: Path) -> None:
    """A workbook opened but with the addressed cell empty is a fail that names it —
    the silent-green case the gate exists to refuse, surfaced through the tool."""
    make_workbook(tmp_path / "book.xlsx", {"A1": "header"})  # B2 is empty
    runner = runner_for(tmp_path)
    args = {
        "workbook": "book.xlsx",
        "default_tolerance": {"abs": 0.01},
        "expectations": [{"cell": "B2", "expected": 5.0, "unit": "MWh"}],
    }
    summary = summary_of(await handle_office_reconcile(runner, args))
    assert summary.risk in ("high", "blocked")
    assert summary.passed == 0


async def test_a_malformed_request_is_a_tool_error_not_an_exception(tmp_path: Path) -> None:
    """A spec missing its required default_tolerance is the agent's own mistake — it
    comes back as a readable tool error the agent fixes, within the byte budget."""
    runner = runner_for(tmp_path)
    result = await handle_office_reconcile(runner, {"workbook": "book.xlsx"})
    assert result.get("is_error") is True
    text = result_text(result)
    assert "fix and retry" in text.lower()
    assert len(text.encode()) <= OFFICE_RECONCILE.max_result_bytes


async def test_the_worst_list_is_capped_and_states_what_it_withheld(tmp_path: Path) -> None:
    """More than RECONCILE_WORST_N mismatches: the worst list is bounded, and the
    summary states the withheld count so the cap is never read as 'that was all'
    (AXI shape 1)."""
    n = RECONCILE_WORST_N + 4
    cells = {f"B{i}": float(1000 * i) for i in range(1, n + 1)}  # all wildly wrong
    make_workbook(tmp_path / "book.xlsx", cells)
    runner = runner_for(tmp_path)
    args = {
        "workbook": "book.xlsx",
        "default_tolerance": {"abs": 0.5},
        "expectations": [
            {"cell": f"B{i}", "expected": 0.0, "unit": "MWh"} for i in range(1, n + 1)
        ],
    }
    summary = summary_of(await handle_office_reconcile(runner, args))
    assert summary.risk == "high"
    assert summary.failed == n
    assert len(summary.worst) == RECONCILE_WORST_N
    assert summary.withheld == n - RECONCILE_WORST_N
    assert f"{summary.withheld} more" in summary.next_step


async def test_an_offset_bearing_expectation_comes_back_as_a_teaching_fail(tmp_path: Path) -> None:
    """The DST-fold defeat, through the tool an agent actually calls.

    An agent whose numbers came out of pandas writes the fall-back hour the idiomatic
    way — ``2024-10-27T02:00:00+02:00`` and ``+01:00``, the only thing that tells the
    two occurrences apart. The gate used to strip the offset, collapsing both to one
    lookup key, so the second was scored against the first one's row and the tool
    answered ``pass``. It must be an honest fail, and the reason the agent reads has to
    survive the tool's own reason clip with the fix still in it — a lesson cut off at
    140 bytes is not a lesson.
    """
    make_workbook(tmp_path / "prices.xlsx", {"A1": "timestamp", "B1": "price"})
    wb = openpyxl.load_workbook(tmp_path / "prices.xlsx")
    ws = wb["Sheet1"]
    ws["A2"], ws["B2"] = datetime(2024, 10, 27, 2), 20.0  # the CEST 02:00
    ws["A3"], ws["B3"] = datetime(2024, 10, 27, 2), 21.0  # the CET 02:00
    wb.save(tmp_path / "prices.xlsx")
    runner = runner_for(tmp_path)
    args = {
        "workbook": "prices.xlsx",
        "timezone": "Europe/Oslo",
        "default_tolerance": {"abs": 0.001},
        "time_index": {
            "timestamp_column": "A",
            "value_column": "B",
            "value_unit": "EUR/MWh",
            "sheet": "Sheet1",
            "expectations": [
                {"timestamp": "2024-10-27T02:00:00+02:00", "expected": 20.0, "unit": "EUR/MWh"},
                {"timestamp": "2024-10-27T02:00:00+01:00", "expected": 20.0, "unit": "EUR/MWh"},
            ],
        },
    }
    result = await handle_office_reconcile(runner, args)
    summary = summary_of(result)
    assert summary.risk == "high", summary.summary
    assert summary.failed == 2
    assert summary.passed == 0
    assert len(summary.worst) == 2
    for mismatch in summary.worst:
        assert mismatch.actual is None  # never scored against the row it did not address
        reason = mismatch.reason or ""
        assert "UTC offset" in reason
        # The clip did not eat the instruction that fixes it.
        assert "fold" in reason
    assert len(result_text(result).encode()) <= OFFICE_RECONCILE.max_result_bytes


async def test_a_non_finite_expected_stays_valid_json(tmp_path: Path) -> None:
    """An agent's tool-call JSON can carry ``"expected": 1e400`` — a valid JSON number
    literal that Python's parser silently overflows to ``inf`` (``ExpectedValue.expected``
    has no ``allow_inf_nan`` guard). That ``inf`` flows through the comparison into the
    mismatch, and ``model_dump_json`` would emit the bare token ``Infinity`` — not valid
    JSON per RFC 8259 — defeating the machine-readable result. The tool must drop the
    non-finite value so the summary parses cleanly with no ``Infinity``/``NaN`` tokens."""
    make_workbook(tmp_path / "book.xlsx", {"B2": 1.0})
    runner = runner_for(tmp_path)
    # Parse the args from raw JSON exactly as the SDK hands a tool call over, so the
    # 1e400 literal overflows to float('inf') the same way it does in production.
    args = json.loads(
        '{"workbook": "book.xlsx", "default_tolerance": {"abs": 0.01}, '
        '"expectations": [{"cell": "B2", "expected": 1e400, "unit": "MWh"}]}'
    )
    result = await handle_office_reconcile(runner, args)
    text = result_text(result)
    # The whole point: a strict JSON parser must accept it, and neither non-standard
    # token may appear anywhere in the body.
    json.loads(text)
    assert "Infinity" not in text
    assert "NaN" not in text
    summary = summary_of(result)
    assert summary.failed == 1
    assert len(summary.worst) == 1
    # The overflowing expected was dropped to None rather than serialized as Infinity.
    assert summary.worst[0].expected is None


# ------------------------------------------------------------------------------ budget


def test_the_tool_is_registered_and_allow_listed() -> None:
    """One place a tool is added: it rides AGENT_TOOLS, so every session kind — the
    same set office_read/office_write get — allows it."""
    assert "mcp__workbench__office_reconcile" in allowed_tool_names("chat")
    assert "mcp__workbench__office_reconcile" in allowed_tool_names("orchestrator")


def test_the_description_fits_the_ceiling() -> None:
    """The tool's own description budget, beside the tool (also swept by
    test_agent_tools' ALL_AGENT_TOOLS iteration)."""
    assert len(OFFICE_RECONCILE.description) <= MAX_DESCRIPTION_CHARS
    assert "\n" not in OFFICE_RECONCILE.description


def test_the_schema_fits_the_ceiling() -> None:
    assert OFFICE_RECONCILE.schema_bytes <= OFFICE_RECONCILE.max_schema_bytes
    assert OFFICE_RECONCILE.input_schema["required"] == ["workbook", "default_tolerance"]


async def test_a_worst_case_result_stays_within_the_byte_budget(tmp_path: Path) -> None:
    """The serialized-result ceiling, sized from the payload next to it: the worst
    case is RECONCILE_WORST_N mismatches, each with a long agent-supplied cell label
    and a long reason, plus a summary and a next-step line. This is compact JSON the
    agent parses, so it is bounded by construction (clipped fields, fixed N), never
    clamped — clamping valid JSON would hand back invalid JSON."""
    # Long labels the agent controls, to push the clip paths and the widest body.
    long_label = "Åsen 2 day-ahead revenue, delivery hour " + "x" * 80
    n = RECONCILE_WORST_N + 3
    cells = {f"B{i}": float(999999 * i) for i in range(1, n + 1)}
    make_workbook(tmp_path / "book.xlsx", cells)
    runner = runner_for(tmp_path)
    args = {
        "workbook": "book.xlsx",
        "default_tolerance": {"abs": 0.5},
        "expectations": [
            {
                "cell": f"{long_label}!B{i}",
                "expected": 0.123456789,
                "unit": "kWh",
                "cell_unit": "MWh",
            }
            for i in range(1, n + 1)
        ],
    }
    result = await handle_office_reconcile(runner, args)
    text = result_text(result)
    assert len(text.encode()) <= OFFICE_RECONCILE.max_result_bytes
    # Compact JSON, not pretty, and valid (parseable) — never truncated mid-token.
    assert "\n" not in text
    json.loads(text)
