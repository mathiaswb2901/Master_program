"""The workbook↔code numeric reconciliation gate — M6's first domain check.

Fake-first / no Office: every fixture is a tiny ``.xlsx`` built with openpyxl in the
test, read back with openpyxl, so the whole gate is exercised deterministically on the
CI machine. What is under test is the *domain* behaviour that makes this more than a
spreadsheet diff:

* a matching workbook is ``pass`` and a mismatch beyond tolerance is a ``fail`` that
  derives risk ``high`` — never a silent green;
* a value 1000x off because kWh was read as MWh is **caught**, and a legitimate unit
  conversion is *named*; an incompatible dimension (energy vs power) is a fail;
* a fall-back DST day's 25 hourly rows (two 02:00s) align by local wall-clock time, not
  by row position;
* a missing/empty cell and an unreadable/unfound workbook are explicit fails that name
  the reason;
* the full comparison table is stored as a bounded payload and reachable through the
  service's payload accessor — it never rides the result;
* the check registers into the production wiring and ``POST /api/validation/run`` with
  its id runs a real reconciliation.
"""

from datetime import datetime
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

import openpyxl
import pytest
from fastapi.testclient import TestClient

from workbench_server.config import Settings
from workbench_server.main import create_app
from workbench_server.models.reconciliation import (
    ExpectedValue,
    ReconciliationReport,
    ReconciliationSpec,
    TimeExpectation,
    TimeIndexSpec,
    Tolerance,
)
from workbench_server.models.validation import ValidationResult, ValidationSpec, ValidationSubject
from workbench_server.services.event_bus import EventBus

# `_align_time_rows` and `_OFFSET_REFUSAL` are reached by name on purpose: the former
# is what consumes `WorkbookReader.column_pairs`, and is the only way to exercise a
# tz-aware datetime *object* (openpyxl refuses to write one into an .xlsx); the latter
# is the refusal text whose length has to fit the agent tool's reason clip.
from workbench_server.services.reconciliation import (
    _OFFSET_REFUSAL,
    MAX_DUPLICATE_LINES,
    CrossCurrency,
    ReconciliationCheck,
    UnitMismatch,
    _align_time_rows,
    convert,
    within_tolerance,
)
from workbench_server.services.validation import ValidationService

# --------------------------------------------------------------------------- helpers


def make_workbook(path: Path, cells: dict[str, Any], sheet: str = "Sheet1") -> None:
    """A tiny workbook with the given A1 cells populated."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet
    for addr, value in cells.items():
        ws[addr] = value
    wb.save(path)


def make_dst_workbook(path: Path) -> None:
    """A price sheet for the Europe/Oslo fall-back day 2024-10-27: 25 hourly rows,
    with 02:00 appearing twice (the CEST hour then the CET hour).

    The value at each hour is deliberately its hour number, except the two 02:00s
    which carry distinct values (20.0 then 21.0) so a fold mix-up is visible. The
    05:00 row sits at *data index 6* (two 02:00s pushed it down), so any positional
    join would pair the 05:00 expectation against the 04:00 row and fail.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Prices"
    ws["A1"] = "timestamp"
    ws["B1"] = "price_eur_mwh"
    rows: list[tuple[datetime, float]] = []
    rows.append((datetime(2024, 10, 27, 0), 0.0))
    rows.append((datetime(2024, 10, 27, 1), 1.0))
    rows.append((datetime(2024, 10, 27, 2), 20.0))  # fold 0 (CEST)
    rows.append((datetime(2024, 10, 27, 2), 21.0))  # fold 1 (CET)
    for hour in range(3, 24):
        rows.append((datetime(2024, 10, 27, hour), float(hour)))
    for offset, (ts, value) in enumerate(rows):
        r = 2 + offset
        ws[f"A{r}"] = ts
        ws[f"B{r}"] = value
    wb.save(path)


def make_duplicated_row_workbook(path: Path) -> None:
    """An *ordinary data-entry error*: a plain June day whose 02:00 row was pasted
    twice. June has no DST transition in Europe/Oslo, so the second 02:00 is not a
    repeated hour — it is a duplicated row, and a validation gate must say so.

    The two copies carry different values (20.0 then 21.0) so that a gate which
    silently treated them as fold 0 / fold 1 would happily *match* a fold-1
    expectation of 21.0 — i.e. bless corrupt input.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Prices"
    ws["A1"] = "timestamp"
    ws["B1"] = "price_eur_mwh"
    rows: list[tuple[datetime, float]] = [
        (datetime(2024, 6, 15, 0), 0.0),
        (datetime(2024, 6, 15, 1), 1.0),
        (datetime(2024, 6, 15, 2), 20.0),
        (datetime(2024, 6, 15, 2), 21.0),  # the pasted duplicate
        (datetime(2024, 6, 15, 3), 3.0),
    ]
    for offset, (ts, value) in enumerate(rows):
        r = 2 + offset
        ws[f"A{r}"] = ts
        ws[f"B{r}"] = value
    wb.save(path)


async def run_recon(
    root: Path, spec: ReconciliationSpec
) -> tuple[ValidationService, ValidationResult]:
    """Register the real check on a service and run a reconciliation end-to-end."""
    service = ValidationService(root, EventBus())
    service.register(ReconciliationCheck())
    subject = ValidationSubject(kind="file", ref=spec.workbook, label=spec.workbook)
    vspec = ValidationSpec(subject=subject, checks=["reconciliation"], params=spec.model_dump())
    result = await service.run(vspec)
    return service, result


def report_of(service: ValidationService, result: ValidationResult) -> ReconciliationReport:
    """The stored table behind the grouped evidence line."""
    ref = result.evidence[0].payload_ref
    assert ref is not None
    payload = service.payload("numeric", ref)
    assert isinstance(payload, ReconciliationReport)
    return payload


# --------------------------------------------------------------- unit conversion core


def test_within_tolerance_is_never_a_naive_equality() -> None:
    tol = Tolerance(abs=0.5)
    assert within_tolerance(10.0, 10.4, tol)
    assert not within_tolerance(10.0, 10.6, tol)
    rel = Tolerance(rel=0.01)
    assert within_tolerance(1000.0, 1009.0, rel)
    assert not within_tolerance(1000.0, 1011.0, rel)


def test_convert_names_a_conversion_and_refuses_incompatible_dimensions() -> None:
    converted, note = convert(5000.0, "kWh", "MWh")
    assert converted == pytest.approx(5.0)
    assert note is not None and "kWh" in note and "MWh" in note
    # identity is silent
    assert convert(5.0, "MWh", "MWh") == (5.0, None)
    # energy vs power cannot be coerced
    with pytest.raises(UnitMismatch):
        convert(5.0, "MW", "MWh")


# --------------------------------------------------------------- the matching cases


@pytest.mark.asyncio
async def test_a_matching_workbook_is_all_pass_and_risk_pass(tmp_path: Path) -> None:
    make_workbook(tmp_path / "book.xlsx", {"D14": 42.5, "E2": 1000.0})
    spec = ReconciliationSpec(
        workbook="book.xlsx",
        default_tolerance=Tolerance(abs=0.001),
        expectations=[
            ExpectedValue(cell="Sheet1!D14", expected=42.5, unit="MWh"),
            ExpectedValue(cell="E2", expected=1000.0, unit="EUR"),
        ],
    )
    service, result = await run_recon(tmp_path, spec)
    assert result.risk == "pass"
    assert result.evidence[0].outcome == "pass"
    assert "reconcile within tolerance" in result.evidence[0].detail
    report = report_of(service, result)
    assert (report.matched, report.mismatched, report.total) == (2, 0, 2)


@pytest.mark.asyncio
async def test_within_tolerance_passes(tmp_path: Path) -> None:
    make_workbook(tmp_path / "book.xlsx", {"D14": 42.5004})
    spec = ReconciliationSpec(
        workbook="book.xlsx",
        default_tolerance=Tolerance(abs=0.001),
        expectations=[ExpectedValue(cell="D14", expected=42.5, unit="MWh")],
    )
    _, result = await run_recon(tmp_path, spec)
    assert result.risk == "pass"


@pytest.mark.asyncio
async def test_a_mismatch_beyond_tolerance_is_a_fail_and_risk_high(tmp_path: Path) -> None:
    make_workbook(tmp_path / "book.xlsx", {"D14": 50.0})
    spec = ReconciliationSpec(
        workbook="book.xlsx",
        default_tolerance=Tolerance(abs=0.001, rel=0.001),
        expectations=[ExpectedValue(cell="D14", expected=42.5, unit="MWh")],
    )
    service, result = await run_recon(tmp_path, spec)
    assert result.risk == "high"
    assert result.evidence[0].outcome == "fail"
    # AXI shape 3: the detail ends by naming the workbook and the first bad cell.
    assert "book.xlsx" in result.evidence[0].detail and "D14" in result.evidence[0].detail
    row = report_of(service, result).comparisons[0]
    assert row.delta == pytest.approx(7.5)
    assert "outside tolerance" in (row.reason or "")


@pytest.mark.asyncio
async def test_per_cell_tolerance_overrides_the_default(tmp_path: Path) -> None:
    make_workbook(tmp_path / "book.xlsx", {"D14": 43.0})
    spec = ReconciliationSpec(
        workbook="book.xlsx",
        default_tolerance=Tolerance(abs=0.001),
        per_cell_tolerance={"D14": Tolerance(abs=1.0)},
        expectations=[ExpectedValue(cell="D14", expected=42.5, unit="MWh")],
    )
    _, result = await run_recon(tmp_path, spec)
    assert result.risk == "pass"  # would fail under the 0.001 default


# --------------------------------------------------------------- the unit failure modes


@pytest.mark.asyncio
async def test_a_kwh_value_against_an_mwh_expectation_is_caught_not_silent(
    tmp_path: Path,
) -> None:
    """The whole point of the gate: a cell holding 5000 (kWh magnitude) checked
    against a 5.0 MWh expectation is a 1000x error, and it must be a fail — not a
    green tick — when the analyst did not declare the cell was in kWh."""
    make_workbook(tmp_path / "book.xlsx", {"D14": 5000.0})
    spec = ReconciliationSpec(
        workbook="book.xlsx",
        default_tolerance=Tolerance(rel=0.01),
        expectations=[ExpectedValue(cell="D14", expected=5.0, unit="MWh")],
    )
    service, result = await run_recon(tmp_path, spec)
    assert result.risk == "high"
    row = report_of(service, result).comparisons[0]
    assert row.outcome == "fail"
    assert row.delta == pytest.approx(4995.0)


@pytest.mark.asyncio
async def test_a_declared_kwh_cell_converts_and_names_the_conversion(tmp_path: Path) -> None:
    """The legitimate case: the analyst declares the cell is in kWh, so 5000 kWh
    converts to 5 MWh, matches, and the conversion is named in the evidence so the
    1000x can never hide inside the pass."""
    make_workbook(tmp_path / "book.xlsx", {"D14": 5000.0})
    spec = ReconciliationSpec(
        workbook="book.xlsx",
        default_tolerance=Tolerance(rel=0.001),
        expectations=[ExpectedValue(cell="D14", expected=5.0, unit="MWh", cell_unit="kWh")],
    )
    service, result = await run_recon(tmp_path, spec)
    assert result.risk == "pass"
    row = report_of(service, result).comparisons[0]
    assert row.outcome == "pass"
    assert "kWh" in (row.reason or "") and "MWh" in (row.reason or "")


@pytest.mark.asyncio
async def test_a_declared_kwh_price_converts_and_names_the_conversion(tmp_path: Path) -> None:
    """The price sibling of the energy conversion: a cell declared in EUR/kWh checked
    against an EUR/MWh expectation is a x1000, and the legitimate declared case must
    convert (0.05 EUR/kWh → 50 EUR/MWh), match, and *name* the conversion — so a price
    unit slip can no more hide inside a pass than an energy one can."""
    make_workbook(tmp_path / "book.xlsx", {"D14": 0.05})
    spec = ReconciliationSpec(
        workbook="book.xlsx",
        default_tolerance=Tolerance(rel=0.001),
        expectations=[
            ExpectedValue(cell="D14", expected=50.0, unit="EUR/MWh", cell_unit="EUR/kWh")
        ],
    )
    service, result = await run_recon(tmp_path, spec)
    assert result.risk == "pass"
    row = report_of(service, result).comparisons[0]
    assert row.outcome == "pass"
    assert row.actual == pytest.approx(50.0)  # 0.05 * 1000
    assert "EUR/kWh" in (row.reason or "") and "EUR/MWh" in (row.reason or "")


@pytest.mark.asyncio
async def test_an_undeclared_eur_kwh_price_is_caught_not_silent(tmp_path: Path) -> None:
    """And the failure sibling: a cell holding 0.05 (EUR/kWh magnitude) checked against
    a 50.0 EUR/MWh expectation with *no* cell_unit declared is a 1000x error, and must
    be a fail rather than a silent mismatch swallowed by tolerance."""
    make_workbook(tmp_path / "book.xlsx", {"D14": 0.05})
    spec = ReconciliationSpec(
        workbook="book.xlsx",
        default_tolerance=Tolerance(rel=0.01),
        expectations=[ExpectedValue(cell="D14", expected=50.0, unit="EUR/MWh")],
    )
    service, result = await run_recon(tmp_path, spec)
    assert result.risk == "high"
    assert report_of(service, result).comparisons[0].outcome == "fail"


@pytest.mark.asyncio
async def test_incompatible_dimensions_are_a_fail(tmp_path: Path) -> None:
    make_workbook(tmp_path / "book.xlsx", {"D14": 42.5})
    spec = ReconciliationSpec(
        workbook="book.xlsx",
        default_tolerance=Tolerance(abs=0.001),
        expectations=[ExpectedValue(cell="D14", expected=42.5, unit="MWh", cell_unit="MW")],
    )
    service, result = await run_recon(tmp_path, spec)
    assert result.risk == "high"
    row = report_of(service, result).comparisons[0]
    assert row.outcome == "fail"
    assert "unit mismatch" in (row.reason or "")


# --------------------------------------------------------------- the DST alignment


@pytest.mark.asyncio
async def test_a_fall_back_dst_day_aligns_by_wall_clock_not_position(tmp_path: Path) -> None:
    make_dst_workbook(tmp_path / "prices.xlsx")
    # Expectations given deliberately out of order, and including both 02:00 folds.
    spec = ReconciliationSpec(
        workbook="prices.xlsx",
        timezone="Europe/Oslo",
        default_tolerance=Tolerance(abs=0.001),
        time_index=TimeIndexSpec(
            timestamp_column="A",
            value_column="B",
            value_unit="EUR/MWh",
            sheet="Prices",
            expectations=[
                TimeExpectation(timestamp="2024-10-27T05:00:00", expected=5.0, unit="EUR/MWh"),
                TimeExpectation(
                    timestamp="2024-10-27T02:00:00", expected=21.0, unit="EUR/MWh", fold=1
                ),
                TimeExpectation(
                    timestamp="2024-10-27T02:00:00", expected=20.0, unit="EUR/MWh", fold=0
                ),
                TimeExpectation(timestamp="2024-10-27T23:00:00", expected=23.0, unit="EUR/MWh"),
            ],
        ),
    )
    service, result = await run_recon(tmp_path, spec)
    assert result.risk == "pass", result.evidence[0].detail
    report = report_of(service, result)
    assert report.total == 4
    # The two 02:00 folds resolved to their *distinct* values, proving the join is
    # not positional.
    by_cell = {c.cell: c for c in report.comparisons}
    assert by_cell["2024-10-27T02:00:00"].actual == pytest.approx(20.0)  # fold 0
    assert by_cell["2024-10-27T02:00:00 (fold 1)"].actual == pytest.approx(21.0)
    assert by_cell["2024-10-27T05:00:00"].actual == pytest.approx(5.0)


@pytest.mark.asyncio
async def test_a_time_row_with_no_matching_wall_clock_is_a_fail(tmp_path: Path) -> None:
    make_dst_workbook(tmp_path / "prices.xlsx")
    spec = ReconciliationSpec(
        workbook="prices.xlsx",
        timezone="Europe/Oslo",
        default_tolerance=Tolerance(abs=0.001),
        time_index=TimeIndexSpec(
            timestamp_column="A",
            value_column="B",
            sheet="Prices",
            expectations=[
                # a third 02:00 (fold 2) does not exist in the workbook
                TimeExpectation(
                    timestamp="2024-10-27T02:00:00", expected=99.0, unit="EUR/MWh", fold=2
                ),
            ],
        ),
    )
    service, result = await run_recon(tmp_path, spec)
    assert result.risk == "high"
    assert "alignment gap" in (report_of(service, result).comparisons[0].reason or "")


# ------------------------ the naive-local contract: an offset is refused, not stripped


def test_offset_bearing_fold_expectations_are_named_fails_not_a_silent_pass(
    tmp_path: Path,
) -> None:
    """The bug this test was written for, driven the way a user hits it: POST
    /api/validation/run with the spec an agent submits.

    ``2024-10-27T02:00:00+02:00`` and ``2024-10-27T02:00:00+01:00`` are the two
    occurrences of the Europe/Oslo fall-back hour, told apart *only* by their UTC
    offset — the idiomatic, DST-safe way a pandas or Python export writes them. The
    gate used to strip the offset, so both collapsed to the same naive wall clock with
    ``fold`` defaulting to 0, and therefore to the **same lookup key**: the genuine
    fold-1 row (21.0) was never consulted and the second expectation was silently
    scored against the fold-0 value (20.0).

    Both expectations here claim 20.0, so before the fix this run came back ``risk
    pass`` — a green light on numbers whose CET hour is wrong, the fold gate defeated
    on the one day it exists to protect. Each must now be a *named* per-row fail that
    teaches the fix, and neither may be scored against a row it did not address.
    """
    make_dst_workbook(tmp_path / "prices.xlsx")
    params = {
        "workbook": "prices.xlsx",
        "timezone": "Europe/Oslo",
        "default_tolerance": {"abs": 0.001},
        "time_index": {
            "timestamp_column": "A",
            "value_column": "B",
            "value_unit": "EUR/MWh",
            "sheet": "Prices",
            "expectations": [
                {"timestamp": "2024-10-27T02:00:00+02:00", "expected": 20.0, "unit": "EUR/MWh"},
                {"timestamp": "2024-10-27T02:00:00+01:00", "expected": 20.0, "unit": "EUR/MWh"},
            ],
        },
    }
    app = create_app(Settings(workspace_root=tmp_path, fake_agent=True))
    with TestClient(app) as client:
        posted = client.post(
            "/api/validation/run",
            json={
                "subject": {"kind": "file", "ref": "prices.xlsx", "label": "prices.xlsx"},
                "checks": ["reconciliation"],
                "params": params,
            },
        ).json()
        assert posted["risk"] == "high", posted["evidence"][0]["detail"]
        ref = posted["evidence"][0]["payload_ref"]
        assert ref is not None
        report = app.state.validation.payload("numeric", ref)
    assert isinstance(report, ReconciliationReport)
    rows = {c.cell: c for c in report.comparisons}
    assert set(rows) == {"2024-10-27T02:00:00+02:00", "2024-10-27T02:00:00+01:00"}
    for row in rows.values():
        assert row.outcome == "fail"
        # Never scored against the fold-0 row it did not address — the silent pass.
        assert row.actual is None
        reason = row.reason or ""
        assert "UTC offset" in reason
        # It teaches the fix rather than only refusing.
        assert "fold" in reason and "naive local wall clock" in reason
        # And it is not mistaken for a malformed string: the value parses fine.
        assert "unparseable" not in reason


@pytest.mark.asyncio
async def test_a_utc_z_timestamp_is_refused_by_the_same_rule(tmp_path: Path) -> None:
    """``fromisoformat`` accepts a trailing ``Z`` on 3.11+, so a UTC-stamped
    expectation would have been stripped to a naive *UTC* wall clock and then compared
    as if it were local — an off-by-the-offset bug that is invisible on any hour, not
    just the fall-back one."""
    make_dst_workbook(tmp_path / "prices.xlsx")
    spec = ReconciliationSpec(
        workbook="prices.xlsx",
        timezone="Europe/Oslo",
        default_tolerance=Tolerance(abs=0.001),
        time_index=TimeIndexSpec(
            timestamp_column="A",
            value_column="B",
            value_unit="EUR/MWh",
            sheet="Prices",
            expectations=[
                TimeExpectation(timestamp="2024-10-27T05:00:00Z", expected=5.0, unit="EUR/MWh")
            ],
        ),
    )
    service, result = await run_recon(tmp_path, spec)
    assert result.risk == "high"
    row = report_of(service, result).comparisons[0]
    assert row.outcome == "fail"
    assert "UTC offset" in (row.reason or "")


@pytest.mark.asyncio
async def test_a_genuinely_unparseable_timestamp_still_says_unparseable(tmp_path: Path) -> None:
    """The two refusals stay distinguishable: a string that is not a timestamp at all
    gets the malformed-input message, not the offset lesson (which would send the
    analyst looking for an offset that is not there)."""
    make_dst_workbook(tmp_path / "prices.xlsx")
    spec = ReconciliationSpec(
        workbook="prices.xlsx",
        timezone="Europe/Oslo",
        default_tolerance=Tolerance(abs=0.001),
        time_index=TimeIndexSpec(
            timestamp_column="A",
            value_column="B",
            value_unit="EUR/MWh",
            sheet="Prices",
            expectations=[TimeExpectation(timestamp="hour 5", expected=5.0, unit="EUR/MWh")],
        ),
    )
    service, result = await run_recon(tmp_path, spec)
    assert result.risk == "high"
    reason = report_of(service, result).comparisons[0].reason or ""
    assert "unparseable timestamp" in reason
    assert "UTC offset" not in reason


@pytest.mark.asyncio
async def test_an_offset_bearing_timestamp_column_is_named_not_stripped(tmp_path: Path) -> None:
    """The same refusal on the *workbook* side, where the offsets arrive as text —
    what ``df.index.astype(str)`` writes for a tz-aware index (openpyxl cannot store a
    tz-aware datetime at all).

    Stripping them there is the same defect wearing the other hat: with the offset
    gone, which of the two 02:00 rows is fold 0 is decided by **row order alone**, the
    one thing ``_align_time_rows`` promises never to infer a fold from — so this
    descending export would quietly swap the two hours. The rows are therefore not
    indexed, and the finding says so by name: a bare "alignment gap" would send the
    analyst hunting for rows that are sitting right there in the column.
    """
    path = tmp_path / "prices.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Prices"
    ws["A1"] = "timestamp"
    ws["B1"] = "price_eur_mwh"
    # Newest-first, so order and offset disagree about which hour comes first.
    ws["A2"], ws["B2"] = "2024-10-27T02:00:00+01:00", 21.0  # the CET hour
    ws["A3"], ws["B3"] = "2024-10-27T02:00:00+02:00", 20.0  # the CEST hour
    wb.save(path)
    spec = ReconciliationSpec(
        workbook="prices.xlsx",
        timezone="Europe/Oslo",
        default_tolerance=Tolerance(abs=0.001),
        time_index=TimeIndexSpec(
            timestamp_column="A",
            value_column="B",
            value_unit="EUR/MWh",
            sheet="Prices",
            expectations=[
                TimeExpectation(
                    timestamp="2024-10-27T02:00:00", expected=20.0, unit="EUR/MWh", fold=0
                ),
                TimeExpectation(
                    timestamp="2024-10-27T02:00:00", expected=21.0, unit="EUR/MWh", fold=1
                ),
            ],
        ),
    )
    service, result = await run_recon(tmp_path, spec)
    assert result.risk == "high", result.summary
    named = [e for e in result.evidence if "UTC offset" in e.detail]
    assert named, [e.detail for e in result.evidence]
    assert named[0].outcome == "fail"
    detail = named[0].detail
    assert "2 row(s) in column A" in detail
    assert "2024-10-27T02:00:00+01:00" in detail  # the first offender, quoted back
    assert "Europe/Oslo" in detail
    assert "re-run" in detail  # AXI shape 3: what to do next
    # One finding for the whole column, never one per row.
    assert len([e for e in result.evidence if "UTC offset" in e.detail]) == 1
    # The refused rows were not indexed, so neither fold was matched by position.
    rows = {c.cell: c for c in report_of(service, result).comparisons}
    assert rows["2024-10-27T02:00:00"].actual is None
    assert rows["2024-10-27T02:00:00 (fold 1)"].actual is None


def test_a_tz_aware_datetime_from_the_reader_seam_is_refused_too() -> None:
    """The other shape an offset arrives in: a real ``datetime`` with a ``tzinfo``.

    Driven at :func:`_align_time_rows` — the function that consumes
    ``WorkbookReader.column_pairs`` — because it cannot be reached through an
    ``.xlsx``: openpyxl refuses to *write* a tz-aware datetime ("Excel does not
    support timezones in datetimes"). The live-Office COM reader that slots in behind
    the same protocol hands back real values, so this is the branch that guards it.
    """
    zone = ZoneInfo("Europe/Oslo")
    aware = datetime(2024, 10, 27, 2, tzinfo=zone)
    aligned = _align_time_rows([(aware, 20.0), (datetime(2024, 10, 27, 3), 3.0)], zone)
    assert aligned.offset_rows == 1
    assert aligned.offset_sample is not None and "02:00" in aligned.offset_sample
    # The aware row is refused, not stripped into the index; the naive one is kept.
    assert list(aligned.values) == [(datetime(2024, 10, 27, 3), 0)]
    assert aligned.duplicates == {}


def test_the_offset_refusal_survives_the_agent_tools_reason_clip() -> None:
    """``office_reconcile`` clips a comparison's ``reason`` to 140 characters to hold
    its result inside the byte budget, so a lesson longer than that reaches the agent
    with its fix cut off. The refusal is written to fit whole."""
    assert len(_OFFSET_REFUSAL) <= 140
    assert "fold" in _OFFSET_REFUSAL


# ------------------------------------------- duplicated rows are not a silent fold


@pytest.mark.asyncio
async def test_a_duplicated_non_ambiguous_timestamp_is_a_fail_not_a_silent_fold(
    tmp_path: Path,
) -> None:
    """The defect this regression test was written for: a workbook whose 02:00 row on a
    *June* day was pasted twice used to be read as fold 0 / fold 1 — so a fold-1
    expectation MATCHED and the gate reported 'reconciled' on corrupt input, inverting
    the whole promise of a validation gate.

    June is not a DST-transition date in Europe/Oslo, so the second 02:00 is not a
    repeated hour. It must be an explicit fail that names the duplicated timestamp, and
    the fold-1 expectation must NOT match.
    """
    make_duplicated_row_workbook(tmp_path / "prices.xlsx")
    spec = ReconciliationSpec(
        workbook="prices.xlsx",
        timezone="Europe/Oslo",
        default_tolerance=Tolerance(abs=0.001),
        time_index=TimeIndexSpec(
            timestamp_column="A",
            value_column="B",
            value_unit="EUR/MWh",
            sheet="Prices",
            expectations=[
                TimeExpectation(
                    timestamp="2024-06-15T02:00:00", expected=21.0, unit="EUR/MWh", fold=1
                ),
            ],
        ),
    )
    service, result = await run_recon(tmp_path, spec)
    assert result.risk == "high", result.summary
    # The duplicate is named in its own evidence line, and says what to do next.
    dup = [e for e in result.evidence if "duplicate" in e.detail.lower()]
    assert dup, [e.detail for e in result.evidence]
    assert dup[0].outcome == "fail"
    assert "2024-06-15T02:00:00" in dup[0].detail
    assert "Europe/Oslo" in dup[0].detail
    # And the fold-1 expectation did not quietly match the pasted second copy.
    row = report_of(service, result).comparisons[0]
    assert row.outcome == "fail"
    assert row.actual != pytest.approx(21.0)


@pytest.mark.asyncio
async def test_a_fold_1_expectation_on_a_non_ambiguous_date_is_refused(tmp_path: Path) -> None:
    """A fold only exists where the zone actually repeats an hour. Asking for fold 1 on
    an ordinary June date is a spec error, and is refused by name rather than reported
    as a generic 'alignment gap' (which would read as a missing workbook row)."""
    make_workbook(tmp_path / "book.xlsx", {"A1": "timestamp", "B1": "v"})
    wb = openpyxl.load_workbook(tmp_path / "book.xlsx")
    ws = wb["Sheet1"]
    ws["A2"] = datetime(2024, 6, 15, 2)
    ws["B2"] = 20.0
    wb.save(tmp_path / "book.xlsx")
    spec = ReconciliationSpec(
        workbook="book.xlsx",
        timezone="Europe/Oslo",
        default_tolerance=Tolerance(abs=0.001),
        time_index=TimeIndexSpec(
            timestamp_column="A",
            value_column="B",
            value_unit="EUR/MWh",
            sheet="Sheet1",
            expectations=[
                TimeExpectation(
                    timestamp="2024-06-15T02:00:00", expected=20.0, unit="EUR/MWh", fold=1
                ),
            ],
        ),
    )
    service, result = await run_recon(tmp_path, spec)
    assert result.risk == "high"
    row = report_of(service, result).comparisons[0]
    assert row.outcome == "fail"
    reason = row.reason or ""
    assert "not an ambiguous" in reason and "Europe/Oslo" in reason
    assert "alignment gap" not in reason


@pytest.mark.asyncio
async def test_an_over_repeated_fall_back_hour_is_not_called_unambiguous(tmp_path: Path) -> None:
    """A *third* 02:00 on the genuine fall-back day is still a duplicate — but the
    evidence must not claim the timestamp "occurs once", which is false there and
    points the analyst at the wrong fix. Europe/Oslo really does write 2024-10-27
    02:00 twice; it is the third copy that is the paste accident."""
    path = tmp_path / "prices.xlsx"
    make_dst_workbook(path)
    wb = openpyxl.load_workbook(path)
    ws = wb["Prices"]
    ws["A27"] = datetime(2024, 10, 27, 2)  # a third 02:00, appended past the 25 rows
    ws["B27"] = 99.0
    wb.save(path)
    spec = ReconciliationSpec(
        workbook="prices.xlsx",
        timezone="Europe/Oslo",
        default_tolerance=Tolerance(abs=0.001),
        time_index=TimeIndexSpec(
            timestamp_column="A",
            value_column="B",
            value_unit="EUR/MWh",
            sheet="Prices",
            expectations=[
                TimeExpectation(
                    timestamp="2024-10-27T02:00:00", expected=21.0, unit="EUR/MWh", fold=1
                ),
            ],
        ),
    )
    service, result = await run_recon(tmp_path, spec)
    assert result.risk == "high", result.summary
    dup = [e for e in result.evidence if "duplicate" in e.detail.lower()]
    assert dup, [e.detail for e in result.evidence]
    assert dup[0].outcome == "fail"
    assert "2024-10-27T02:00:00" in dup[0].detail
    assert "appears 3 times" in dup[0].detail
    # The false sentence the first cut of this fix would have emitted.
    assert "not an ambiguous local time" not in dup[0].detail
    assert "twice" in dup[0].detail
    # The two genuine folds still aligned, so fold 1 matched its real CET value.
    row = report_of(service, result).comparisons[0]
    assert row.outcome == "pass"
    assert row.actual == pytest.approx(21.0)


@pytest.mark.asyncio
async def test_many_duplicated_timestamps_are_bounded_and_say_how_many_were_withheld(
    tmp_path: Path,
) -> None:
    """A paste accident can duplicate thousands of rows. One fail line each would bury
    the verdict and blow the result's token budget, so the named lines are capped and a
    roll-up states the total and what to do — never a silent truncation (AXI shape 1)."""
    path = tmp_path / "prices.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Prices"
    ws["A1"] = "timestamp"
    ws["B1"] = "price"
    row = 2
    for hour in range(24):  # every hour of an ordinary June day, written twice
        for value in (float(hour), float(hour) + 0.5):
            ws[f"A{row}"] = datetime(2024, 6, 15, hour)
            ws[f"B{row}"] = value
            row += 1
    wb.save(path)
    spec = ReconciliationSpec(
        workbook="prices.xlsx",
        timezone="Europe/Oslo",
        default_tolerance=Tolerance(abs=0.001),
        time_index=TimeIndexSpec(
            timestamp_column="A",
            value_column="B",
            value_unit="EUR/MWh",
            sheet="Prices",
            expectations=[
                TimeExpectation(timestamp="2024-06-15T00:00:00", expected=0.0, unit="EUR/MWh"),
            ],
        ),
    )
    _, result = await run_recon(tmp_path, spec)
    assert result.risk == "high"
    named = [e for e in result.evidence if e.label.startswith("duplicate timestamp row (")]
    assert len(named) == MAX_DUPLICATE_LINES
    rollup = [e for e in result.evidence if "more (" in e.label]
    assert len(rollup) == 1
    assert rollup[0].outcome == "fail"
    assert "24 distinct timestamps are duplicated" in rollup[0].detail
    assert f"{24 - MAX_DUPLICATE_LINES} more" in rollup[0].detail


# ------------------------------------------------------------- Nordic currencies


def test_nordic_currency_units_are_first_class() -> None:
    """The beachhead persona trades NO/SE/DK zones, where models are denominated in
    NOK/SEK/DKK per MWh. Within-currency scaling converts and is named, exactly like
    the EUR path."""
    for currency in ("NOK", "SEK", "DKK"):
        converted, note = convert(0.5, f"{currency}/kWh", f"{currency}/MWh")
        assert converted == pytest.approx(500.0)
        assert note is not None and f"{currency}/kWh" in note and f"{currency}/MWh" in note


def test_cross_currency_is_refused_and_no_fx_rate_is_invented() -> None:
    with pytest.raises(CrossCurrency) as caught:
        convert(100.0, "NOK/MWh", "EUR/MWh")
    message = str(caught.value)
    assert "NOK" in message and "EUR" in message
    assert "FX" in message
    # It is still a UnitMismatch, so every existing caller stays safe.
    assert isinstance(caught.value, UnitMismatch)


@pytest.mark.asyncio
async def test_a_declared_nok_kwh_price_converts_and_names_the_conversion(tmp_path: Path) -> None:
    make_workbook(tmp_path / "book.xlsx", {"D14": 0.5})
    spec = ReconciliationSpec(
        workbook="book.xlsx",
        default_tolerance=Tolerance(rel=0.001),
        expectations=[
            ExpectedValue(cell="D14", expected=500.0, unit="NOK/MWh", cell_unit="NOK/kWh")
        ],
    )
    service, result = await run_recon(tmp_path, spec)
    assert result.risk == "pass", result.evidence[0].detail
    row = report_of(service, result).comparisons[0]
    assert row.actual == pytest.approx(500.0)
    assert "NOK/kWh" in (row.reason or "") and "NOK/MWh" in (row.reason or "")


@pytest.mark.asyncio
async def test_a_cross_currency_comparison_is_a_principled_refusal(tmp_path: Path) -> None:
    """NOK/MWh against an EUR/MWh expectation is not an 'unknown unit' — the units are
    both known and both prices. It is a refusal to invent an FX rate, and the evidence
    has to say that, because the fix (state the FX rate you meant, in your own code) is
    different from the fix for a typo'd unit."""
    make_workbook(tmp_path / "book.xlsx", {"D14": 1200.0})
    spec = ReconciliationSpec(
        workbook="book.xlsx",
        default_tolerance=Tolerance(rel=0.01),
        expectations=[
            ExpectedValue(cell="D14", expected=100.0, unit="EUR/MWh", cell_unit="NOK/MWh")
        ],
    )
    service, result = await run_recon(tmp_path, spec)
    assert result.risk == "high"
    row = report_of(service, result).comparisons[0]
    assert row.outcome == "fail"
    reason = row.reason or ""
    assert "cross-currency" in reason
    assert "NOK" in reason and "EUR" in reason
    assert "unknown unit" not in reason


@pytest.mark.asyncio
async def test_an_actually_unknown_unit_still_says_unknown(tmp_path: Path) -> None:
    make_workbook(tmp_path / "book.xlsx", {"D14": 1.0})
    spec = ReconciliationSpec(
        workbook="book.xlsx",
        default_tolerance=Tolerance(rel=0.01),
        expectations=[ExpectedValue(cell="D14", expected=1.0, unit="MWh", cell_unit="bananas")],
    )
    service, result = await run_recon(tmp_path, spec)
    assert result.risk == "high"
    reason = report_of(service, result).comparisons[0].reason or ""
    assert "unknown unit" in reason
    assert "cross-currency" not in reason


# --------------------------------------------------------------- unreadable / missing


@pytest.mark.asyncio
async def test_an_empty_cell_is_a_fail_with_a_reason(tmp_path: Path) -> None:
    make_workbook(tmp_path / "book.xlsx", {"D14": 42.5})  # D15 left empty
    spec = ReconciliationSpec(
        workbook="book.xlsx",
        default_tolerance=Tolerance(abs=0.001),
        expectations=[
            ExpectedValue(cell="D14", expected=42.5, unit="MWh"),
            ExpectedValue(cell="D15", expected=10.0, unit="MWh"),
        ],
    )
    service, result = await run_recon(tmp_path, spec)
    assert result.risk == "high"
    rows = {c.cell: c for c in report_of(service, result).comparisons}
    assert rows["D15"].outcome == "fail"
    assert "empty cell" in (rows["D15"].reason or "")
    assert rows["D14"].outcome == "pass"


@pytest.mark.asyncio
async def test_a_malformed_cell_address_is_isolated_to_one_fail_row(tmp_path: Path) -> None:
    """A plausible typo — a bare row reference '14' with the column letter dropped —
    must fail *only that row*, not crash the whole check. openpyxl resolves ws['14']
    to a tuple of the row's cells whose `.value` raises AttributeError; the gate turns
    that into one CellComparison fail while D14 and D15 still reconcile and land in the
    stored report (proof the run was not downgraded to a single kind='gate' line)."""
    make_workbook(tmp_path / "book.xlsx", {"D14": 42.5, "D15": 10.0})
    spec = ReconciliationSpec(
        workbook="book.xlsx",
        default_tolerance=Tolerance(abs=0.001),
        expectations=[
            ExpectedValue(cell="D14", expected=42.5, unit="MWh"),
            ExpectedValue(cell="14", expected=1.0, unit="MWh"),  # column letter dropped
            ExpectedValue(cell="D15", expected=10.0, unit="MWh"),
        ],
    )
    service, result = await run_recon(tmp_path, spec)
    assert result.risk == "high"
    assert result.evidence[0].kind == "numeric"  # not "gate": the report was still built
    rows = {c.cell: c for c in report_of(service, result).comparisons}
    assert rows["14"].outcome == "fail"
    assert "cannot read 14" in (rows["14"].reason or "")
    assert rows["D14"].outcome == "pass" and rows["D15"].outcome == "pass"


@pytest.mark.asyncio
async def test_an_unknown_time_index_sheet_fails_only_its_rows(tmp_path: Path) -> None:
    """A mistyped TimeIndexSpec.sheet must isolate to the time rows: reading the
    columns off a non-existent sheet raises KeyError from the reader, and the gate
    turns each time expectation into a fail while the direct-cell comparison in the
    same run still lands in the stored report."""
    make_workbook(tmp_path / "book.xlsx", {"D14": 42.5})
    spec = ReconciliationSpec(
        workbook="book.xlsx",
        timezone="Europe/Oslo",
        default_tolerance=Tolerance(abs=0.001),
        expectations=[ExpectedValue(cell="D14", expected=42.5, unit="MWh")],
        time_index=TimeIndexSpec(
            timestamp_column="A",
            value_column="B",
            sheet="NoSuchSheet",
            expectations=[
                TimeExpectation(timestamp="2024-10-27T05:00:00", expected=5.0, unit="EUR/MWh"),
            ],
        ),
    )
    service, result = await run_recon(tmp_path, spec)
    assert result.risk == "high"
    assert result.evidence[0].kind == "numeric"  # not downgraded to a gate line
    rows = {c.cell: c for c in report_of(service, result).comparisons}
    assert rows["D14"].outcome == "pass"  # the valid direct comparison survived
    bad = rows["2024-10-27T05:00:00"]
    assert bad.outcome == "fail"
    assert "time-index columns" in (bad.reason or "")


@pytest.mark.asyncio
async def test_a_workbook_with_no_cached_values_is_blocked_style_fail(tmp_path: Path) -> None:
    make_workbook(tmp_path / "book.xlsx", {"A1": "header"})  # addressed cells all empty
    spec = ReconciliationSpec(
        workbook="book.xlsx",
        default_tolerance=Tolerance(abs=0.001),
        expectations=[ExpectedValue(cell="D14", expected=42.5, unit="MWh")],
    )
    _, result = await run_recon(tmp_path, spec)
    assert result.risk == "high"
    assert "open and save" in result.evidence[0].detail.lower()


@pytest.mark.asyncio
async def test_a_missing_workbook_is_an_explicit_fail(tmp_path: Path) -> None:
    spec = ReconciliationSpec(
        workbook="nope.xlsx",
        default_tolerance=Tolerance(abs=0.001),
        expectations=[ExpectedValue(cell="D14", expected=1.0, unit="MWh")],
    )
    _, result = await run_recon(tmp_path, spec)
    assert result.risk == "high"
    assert "not found" in result.evidence[0].detail


@pytest.mark.asyncio
async def test_a_workbook_path_escaping_the_workspace_is_refused(tmp_path: Path) -> None:
    spec = ReconciliationSpec(
        workbook="../secret.xlsx",
        default_tolerance=Tolerance(abs=0.001),
        expectations=[ExpectedValue(cell="D14", expected=1.0, unit="MWh")],
    )
    _, result = await run_recon(tmp_path, spec)
    assert result.risk == "high"
    assert "escapes the workspace" in result.evidence[0].detail


# --------------------------------------------------------------- the payload seam


@pytest.mark.asyncio
async def test_the_full_table_is_stored_as_a_payload_not_inlined(tmp_path: Path) -> None:
    make_workbook(tmp_path / "book.xlsx", {"D14": 42.5, "D15": 10.0})
    spec = ReconciliationSpec(
        workbook="book.xlsx",
        default_tolerance=Tolerance(abs=0.001),
        expectations=[
            ExpectedValue(cell="D14", expected=42.5, unit="MWh", label="revenue"),
            ExpectedValue(cell="D15", expected=10.0, unit="MWh"),
        ],
    )
    service, result = await run_recon(tmp_path, spec)
    # One grouped evidence line; the table is behind its payload_ref, never inline.
    assert len(result.evidence) == 1
    report = report_of(service, result)
    assert report.total == 2
    assert {c.cell for c in report.comparisons} == {"D14", "D15"}
    assert report.comparisons[0].label == "revenue" or report.comparisons[1].label == "revenue"


# --------------------------------------------------------------- registration wiring


def test_the_check_is_registered_in_the_production_wiring(tmp_path: Path) -> None:
    """POST /api/validation/run with the check id runs a real reconciliation through
    the app create_app wires up — proof the additive registration landed."""
    make_workbook(tmp_path / "book.xlsx", {"D14": 42.5})
    spec = ReconciliationSpec(
        workbook="book.xlsx",
        default_tolerance=Tolerance(abs=0.001),
        expectations=[ExpectedValue(cell="D14", expected=42.5, unit="MWh")],
    )
    app = create_app(Settings(workspace_root=tmp_path, fake_agent=True))
    with TestClient(app) as client:
        body = {
            "subject": {"kind": "file", "ref": "book.xlsx", "label": "book.xlsx"},
            "checks": ["reconciliation"],
            "params": spec.model_dump(),
        }
        posted = client.post("/api/validation/run", json=body).json()
    assert posted["risk"] == "pass"  # not "high"/"not registered" — the check ran
    assert posted["evidence"][0]["kind"] == "numeric"
    assert posted["evidence"][0]["payload_ref"] is not None


def test_a_mismatch_through_the_endpoint_is_high(tmp_path: Path) -> None:
    make_workbook(tmp_path / "book.xlsx", {"D14": 99.0})
    spec = ReconciliationSpec(
        workbook="book.xlsx",
        default_tolerance=Tolerance(rel=0.001),
        expectations=[ExpectedValue(cell="D14", expected=42.5, unit="MWh")],
    )
    app = create_app(Settings(workspace_root=tmp_path, fake_agent=True))
    with TestClient(app) as client:
        body = {
            "subject": {"kind": "file", "ref": "book.xlsx", "label": "book.xlsx"},
            "checks": ["reconciliation"],
            "params": spec.model_dump(),
        }
        posted = client.post("/api/validation/run", json=body).json()
    assert posted["risk"] == "high"
