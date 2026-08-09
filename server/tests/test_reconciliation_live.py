"""The live-reconcile front gate — **no false PASS**.

The defect this suite pins is the one the whole M6 milestone exists to refuse,
arriving through the milestone's own front door: a workbook docked in a panel,
open in a real Excel, with an hour of unsaved edits in it. The gate reads the
``.xlsx`` on disk and reports **pass** — about numbers that are not the numbers.
Measured against a real Excel on this machine
(``scripts/dev/probe_live_com.py``)::

    live B1 after an edit      = 9999.0      Workbook.Saved = False
    disk B1 (openpyxl, again)  = 1234.5      <-- what a disk gate reconciles
    live C1 (=B1*2)            = 19998.0
    disk C1                    = None        (never opened in Excel)

Everything here runs with **no Office**: the four-row decision table is driven
from a twenty-line stub of the ``LiveWorkbookHost`` seam, and the same table is
then driven end-to-end through the real ``OfficeHostService`` with the fake
backend and fake bridge — the ``WORKBENCH_OFFICE_FAKE=1`` posture, which is what
CI and the E2E journey run.

Four claims, in the order they matter:

1. **A docked, dirty workbook whose live numbers cannot be read is ``blocked``,
   naming the remedy** — never a pass against cached disk, and never a bare
   refusal with no sentence in it.
2. **A docked, dirty workbook that *can* be read is judged on the live numbers**,
   and the evidence says so.
3. **The host is consulted before the file is opened.** Asserted on call order,
   because the failure mode is a later edit quietly reordering two lines into a
   read-then-check.
4. **A tz-aware value from COM is normalised by ``replace(tzinfo=None)`` and
   never by ``astimezone``** — the latter is an hour wrong on 2024-10-27, the one
   day the DST gate exists for.
"""

from collections.abc import Sequence
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

import openpyxl
import pytest

from workbench_server.models.office_bridge import CalculationState, LiveWorkbookStatus
from workbench_server.models.office_host import PanelRect
from workbench_server.models.reconciliation import (
    ExpectedValue,
    ReconciliationReport,
    ReconciliationSpec,
    TimeExpectation,
    TimeIndexSpec,
    Tolerance,
)
from workbench_server.models.validation import ValidationResult, ValidationSpec, ValidationSubject
from workbench_server.services.event_bus import EventBus
from workbench_server.services.office_host.fake_backend import FakeHostBackend
from workbench_server.services.office_host.fake_document_bridge import FakeDocumentBridge
from workbench_server.services.office_host.office_com import naive_local
from workbench_server.services.office_host.service import OfficeHostService
from workbench_server.services.reconciliation import (
    DiskWorkbookReader,
    LiveComWorkbookReader,
    ReconciliationCheck,
    describe_source,
)
from workbench_server.services.validation import ValidationService, derive_risk
from workbench_server.services.workspace import Workspace

RECT = PanelRect(x=10, y=20, width=640, height=480)

#: The Nordic fall-back hour, written twice by ``Europe/Oslo``. Correction 4 of
#: the plan is measured on exactly this value.
FALL_BACK = datetime(2024, 10, 27, 2, 0)


# --------------------------------------------------------------------------- helpers


class StubHost:
    """The ``LiveWorkbookHost`` seam, scripted — the whole four-row table in one
    object, and a record of the order it was asked things in."""

    def __init__(
        self,
        *,
        docked: bool = True,
        saved: bool = True,
        calculation: CalculationState = "done",
        cells: dict[str, Any] | None = None,
        columns: list[tuple[Any, Any]] | None = None,
        live_reads: bool = True,
        status_reads: bool = True,
    ) -> None:
        self.docked = docked
        self.status = LiveWorkbookStatus(saved=saved, calculation=calculation)
        self.cells = cells or {}
        self.columns = columns or []
        self.live_reads = live_reads
        self.status_reads = status_reads
        #: Every seam call, in order — plus "disk" when the openpyxl reader is
        #: constructed (the test patches it in to append).
        self.calls: list[str] = []

    def docked_workbook(self, path: str) -> str | None:
        self.calls.append("docked_workbook")
        return path if self.docked else None

    async def workbook_status(self, path: str) -> LiveWorkbookStatus:
        self.calls.append("workbook_status")
        if not self.status_reads:
            raise RuntimeError("the live state could not be read")
        return self.status

    async def read_workbook_cells(
        self, path: str, sheet: str | None, cells: Sequence[str]
    ) -> list[object]:
        self.calls.append("read_workbook_cells")
        if not self.live_reads:
            raise RuntimeError("reading the live document is not available here")
        return [self.cells.get(address) for address in cells]

    async def read_workbook_columns(
        self,
        path: str,
        sheet: str | None,
        ts_column: str,
        value_column: str,
        start_row: int,
        max_rows: int,
    ) -> list[tuple[object, object]]:
        self.calls.append("read_workbook_columns")
        if not self.live_reads:
            raise RuntimeError("reading the live document is not available here")
        return list(self.columns)


def make_workbook(path: Path, cells: dict[str, Any], sheet: str = "Sheet1") -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet
    for addr, value in cells.items():
        ws[addr] = value
    wb.save(path)


def cell_spec(workbook: str, cell: str, expected: float) -> ReconciliationSpec:
    return ReconciliationSpec(
        workbook=workbook,
        expectations=[ExpectedValue(cell=cell, expected=expected, unit="MWh")],
        default_tolerance=Tolerance(rel=0.001),
    )


async def run_recon(
    root: Path, spec: ReconciliationSpec, host: Any = None
) -> tuple[ValidationService, ValidationResult]:
    service = ValidationService(root, EventBus())
    service.register(ReconciliationCheck(host))
    subject = ValidationSubject(kind="file", ref=spec.workbook, label=spec.workbook)
    vspec = ValidationSpec(subject=subject, checks=["reconciliation"], params=spec.model_dump())
    return service, await service.run(vspec)


def report_of(service: ValidationService, result: ValidationResult) -> ReconciliationReport:
    ref = result.evidence[0].payload_ref
    assert ref is not None
    payload = service.payload("numeric", ref)
    assert isinstance(payload, ReconciliationReport)
    return payload


# ------------------------------------------------------- the four-row decision table


class TestFrontGate:
    """One test per row of the table in ``services/reconciliation.py``."""

    async def test_not_docked_reads_disk_and_says_so(self, tmp_path: Path) -> None:
        make_workbook(tmp_path / "model.xlsx", {"B1": 1234.5})
        host = StubHost(docked=False)
        service, result = await run_recon(tmp_path, cell_spec("model.xlsx", "B1", 1234.5), host)

        assert result.risk == "pass"
        report = report_of(service, result)
        assert report.source.kind == "file"
        assert report.source.mtime is not None
        assert report.source.cached_values is True
        # Nothing about the live seam was reached past the one question.
        assert host.calls == ["docked_workbook"]
        assert "Read from the file on disk" in result.evidence[0].detail

    async def test_docked_and_clean_reads_live_when_it_can(self, tmp_path: Path) -> None:
        make_workbook(tmp_path / "model.xlsx", {"B1": 1234.5})
        host = StubHost(saved=True, cells={"B1": 1234.5})
        service, result = await run_recon(tmp_path, cell_spec("model.xlsx", "B1", 1234.5), host)

        assert result.risk == "pass"
        report = report_of(service, result)
        assert report.source.kind == "live"
        assert report.source.saved is True
        assert report.source.calculation == "done"

    async def test_docked_and_clean_falls_back_to_disk_when_live_refuses(
        self, tmp_path: Path
    ) -> None:
        # The file and the instance agree, so disk is a *correct* answer here
        # rather than a convenient one — and the evidence still names which.
        make_workbook(tmp_path / "model.xlsx", {"B1": 1234.5})
        host = StubHost(saved=True, live_reads=False)
        service, result = await run_recon(tmp_path, cell_spec("model.xlsx", "B1", 1234.5), host)

        assert result.risk == "pass"
        assert report_of(service, result).source.kind == "file"

    async def test_docked_and_dirty_is_judged_on_the_live_numbers(self, tmp_path: Path) -> None:
        # The reproduction, at the seam: disk still holds the old number and the
        # live instance holds the edit. A gate that read the file would pass.
        make_workbook(tmp_path / "model.xlsx", {"B1": 1234.5})
        host = StubHost(saved=False, cells={"B1": 9999.0})
        service, result = await run_recon(tmp_path, cell_spec("model.xlsx", "B1", 1234.5), host)

        assert result.risk == "high"
        report = report_of(service, result)
        assert report.source.kind == "live"
        assert report.source.saved is False
        assert report.comparisons[0].actual == 9999.0
        assert "workbook unsaved" in result.evidence[0].detail

    async def test_docked_and_dirty_with_no_live_read_is_blocked(self, tmp_path: Path) -> None:
        # THE row. The file on disk agrees with the expectation perfectly — so a
        # gate that fell back to it would report a green pass about numbers the
        # user has already changed.
        make_workbook(tmp_path / "model.xlsx", {"B1": 1234.5})
        host = StubHost(saved=False, live_reads=False)
        _, result = await run_recon(tmp_path, cell_spec("model.xlsx", "B1", 1234.5), host)

        assert result.risk == "blocked"
        assert [item.outcome for item in result.evidence] == ["blocked"]
        detail = result.evidence[0].detail
        assert "unsaved changes" in detail
        # A refusal that does not name the way out is not worth having.
        assert "Save the workbook" in detail
        assert "desktop shell" in detail
        # And it carries no comparison table: nothing was read, so pretending
        # there is one to expand would be its own small lie.
        assert result.evidence[0].payload_ref is None

    async def test_a_status_that_cannot_be_read_is_blocked_too(self, tmp_path: Path) -> None:
        # Docked, and we could not learn whether it is clean. There is no
        # conservative disk fallback: the unanswered question *is* "would reading
        # the file be wrong?".
        make_workbook(tmp_path / "model.xlsx", {"B1": 1234.5})
        host = StubHost(status_reads=False)
        _, result = await run_recon(tmp_path, cell_spec("model.xlsx", "B1", 1234.5), host)

        assert result.risk == "blocked"
        assert "cannot tell whether it has unsaved changes" in result.evidence[0].detail

    async def test_mid_calculation_is_skipped_not_pass(self, tmp_path: Path) -> None:
        make_workbook(tmp_path / "model.xlsx", {"B1": 1234.5})
        host = StubHost(calculation="calculating", cells={"B1": 1234.5})
        _, result = await run_recon(tmp_path, cell_spec("model.xlsx", "B1", 1234.5), host)

        # Every comparison matched, and the result is still not green.
        assert [item.outcome for item in result.evidence] == ["pass", "skipped"]
        assert result.risk == "low"
        assert "has not finished calculating" in result.evidence[1].detail
        assert "Re-run in a moment" in result.evidence[1].detail

    async def test_the_host_is_consulted_before_the_file_is_opened(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """Order, not outcome. A later edit that opened the workbook first and
        asked the host afterwards would still pass every test above — and would
        be exactly the read-then-check this gate replaced."""
        make_workbook(tmp_path / "model.xlsx", {"B1": 1234.5})
        host = StubHost(docked=False)

        import workbench_server.services.reconciliation as module

        real = module.DiskWorkbookReader

        class Recording(real):  # type: ignore[misc,valid-type]
            def __init__(self, path: Path, read_at: datetime) -> None:
                host.calls.append("disk")
                super().__init__(path, read_at)

        monkeypatch.setattr(module, "DiskWorkbookReader", Recording)
        await run_recon(tmp_path, cell_spec("model.xlsx", "B1", 1234.5), host)
        assert host.calls == ["docked_workbook", "disk"]


# --------------------------------------------------------- the two readers agree


class TestReadersAgree:
    async def test_same_comparisons_different_provenance(self, tmp_path: Path) -> None:
        """A clean workbook read live and read from disk must produce the *same*
        verdict and a *different* ``ReadSource`` — the whole promise of a seam
        with two implementations."""
        make_workbook(tmp_path / "model.xlsx", {"B1": 1234.5})
        spec = cell_spec("model.xlsx", "B1", 1234.5)

        live_service, live_result = await run_recon(
            tmp_path, spec, StubHost(saved=True, cells={"B1": 1234.5})
        )
        disk_service, disk_result = await run_recon(tmp_path, spec, StubHost(docked=False))

        live = report_of(live_service, live_result)
        disk = report_of(disk_service, disk_result)
        assert live.comparisons == disk.comparisons
        assert live.matched == disk.matched == 1
        assert live.source.kind == "live"
        assert disk.source.kind == "file"

    async def test_disk_reader_reports_a_workbook_with_no_cached_values(
        self, tmp_path: Path
    ) -> None:
        # openpyxl writes formulas, never their results: this workbook has been
        # written by a script and opened by no Excel, which is a different thing
        # from a workbook full of zeros — and the source says which.
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["B1"] = "=1+1"
        wb.save(tmp_path / "model.xlsx")
        reader = DiskWorkbookReader(tmp_path / "model.xlsx", datetime(2026, 8, 9, 14, 0))
        assert reader.cell_value(None, "B1") is None
        assert reader.provenance().cached_values is False
        reader.close()


# ------------------------------------------- correction 4: the tz-aware COM datetime


class TestOffsetNormalisation:
    """``.astimezone()`` is the plausible-looking normalisation that silently
    moves every hourly row of the fall-back day. Measured on this machine::

        live (Range.Value over COM) pywintypes.datetime(2024, 10, 27, 2, 0,
                                        tzinfo=TimeZoneInfo('GMT Standard Time', True))
          .replace(tzinfo=None)     -> 2024-10-27 02:00   correct wall clock
          .astimezone().replace()   -> 2024-10-27 03:00   WRONG BY AN HOUR
    """

    @pytest.mark.parametrize("offset_hours", [0, 1, 2, -5])
    def test_naive_local_keeps_the_wall_clock_whatever_the_offset(self, offset_hours: int) -> None:
        aware = FALL_BACK.replace(tzinfo=timezone(timedelta(hours=offset_hours)))
        assert naive_local(aware) == FALL_BACK
        # …and the wrong answer, spelled out so this fails loudly if anyone
        # "normalises" it instead. Converted into *any* other zone the wall clock
        # moves, which on 2024-10-27 means an hourly series silently re-indexed.
        # Asserted against an explicit zone rather than the machine's, so the
        # claim is about the arithmetic and not about where CI runs.
        elsewhere = timezone(timedelta(hours=offset_hours + 1))
        assert aware.astimezone(elsewhere).replace(tzinfo=None) != FALL_BACK

    def test_non_datetimes_pass_through_untouched(self) -> None:
        assert naive_local(1234.5) == 1234.5
        assert naive_local(None) is None
        assert naive_local("text") == "text"
        assert naive_local(FALL_BACK) == FALL_BACK  # already naive

    async def test_the_live_reader_never_hands_a_tz_aware_value_across_the_seam(self) -> None:
        """The seam, not the transport. ``_as_local_datetime`` refuses an
        offset-bearing value on purpose (#120) — so the live reader has to
        normalise on its own side, or every row of a live read would be refused
        by a rule written for offset-bearing *strings* in a workbook."""
        elsewhere = timezone(timedelta(hours=1))
        spec = ReconciliationSpec(
            workbook="hours.xlsx",
            default_tolerance=Tolerance(rel=0.001),
            timezone="Europe/Oslo",
            time_index=TimeIndexSpec(
                timestamp_column="A",
                value_column="B",
                start_row=2,
                expectations=[
                    TimeExpectation(timestamp="2024-10-27T02:00:00", expected=20.0, fold=0),
                    TimeExpectation(timestamp="2024-10-27T02:00:00", expected=21.0, fold=1),
                ],
            ),
        )
        host = StubHost(
            columns=[
                (FALL_BACK.replace(tzinfo=elsewhere), 20.0),
                (FALL_BACK.replace(tzinfo=elsewhere), 21.0),
            ]
        )
        reader = await LiveComWorkbookReader.load(
            host, "hours.xlsx", spec, host.status, datetime(2026, 8, 9, 14, 0)
        )
        pairs = reader.column_pairs(None, "A", "B", 2)
        assert [ts for ts, _ in pairs] == [FALL_BACK, FALL_BACK]
        assert all(getattr(ts, "tzinfo", None) is None for ts, _ in pairs)

    async def test_both_folds_of_the_fall_back_hour_reconcile_through_the_live_reader(
        self, tmp_path: Path
    ) -> None:
        """The end-to-end version: two rows at the same wall clock, addressed by
        fold, matched against the values the *live* instance holds. A shift of an
        hour here would score fold 1 against the 03:00 row."""
        aware = timezone(timedelta(hours=1))
        make_workbook(tmp_path / "hours.xlsx", {"A1": "ts", "B1": "mwh"})
        spec = ReconciliationSpec(
            workbook="hours.xlsx",
            default_tolerance=Tolerance(rel=0.001),
            timezone="Europe/Oslo",
            time_index=TimeIndexSpec(
                timestamp_column="A",
                value_column="B",
                start_row=2,
                expectations=[
                    TimeExpectation(timestamp="2024-10-27T02:00:00", expected=20.0, fold=0),
                    TimeExpectation(timestamp="2024-10-27T02:00:00", expected=21.0, fold=1),
                    TimeExpectation(timestamp="2024-10-27T03:00:00", expected=3.0),
                ],
            ),
        )
        host = StubHost(
            saved=False,
            columns=[
                (FALL_BACK.replace(tzinfo=aware), 20.0),
                (FALL_BACK.replace(tzinfo=aware), 21.0),
                (datetime(2024, 10, 27, 3).replace(tzinfo=aware), 3.0),
            ],
        )
        service, result = await run_recon(tmp_path, spec, host)
        report = report_of(service, result)
        assert report.matched == 3, [c.reason for c in report.comparisons]
        assert result.risk == "pass"
        assert report.source.kind == "live"


# ------------------------------------------------ the "blocked" outcome, end to end


class TestBlockedOutcome:
    """The member ``CheckOutcome`` did not have, and the four tables that had to
    learn it. Each of these would have been a ``KeyError`` on the first blocked
    run."""

    def test_derive_risk_maps_blocked_to_blocked(self) -> None:
        from workbench_server.models.validation import EvidenceItem

        item = EvidenceItem(kind="numeric", label="x", outcome="blocked", detail="d")
        passing = EvidenceItem(kind="numeric", label="y", outcome="pass", detail="d")
        # Not `low` — the difference between a badge someone scrolls past and a
        # badge that stops them.
        assert derive_risk([item]) == "blocked"
        assert derive_risk([passing, item]) == "blocked"
        assert derive_risk([passing]) == "pass"

    async def test_the_summary_names_the_blocked_count(self, tmp_path: Path) -> None:
        make_workbook(tmp_path / "model.xlsx", {"B1": 1234.5})
        _, result = await run_recon(
            tmp_path,
            cell_spec("model.xlsx", "B1", 1234.5),
            StubHost(saved=False, live_reads=False),
        )
        assert result.summary.startswith("blocked: 1 checks (1 blocked)")

    def test_describe_source_says_which_reader_and_what_made_it_trustworthy(self) -> None:
        from workbench_server.models.reconciliation import ReadSource

        live = describe_source(
            ReadSource(
                kind="live",
                read_at=datetime(2026, 8, 9, 14, 2, 11),
                calculation="done",
                saved=False,
            )
        )
        assert "14:02:11" in live and "calculation done" in live and "workbook unsaved" in live
        disk = describe_source(
            ReadSource(
                kind="file",
                read_at=datetime(2026, 8, 9, 14, 2, 11),
                mtime=datetime(2026, 8, 9, 13, 41, 8),
                cached_values=True,
            )
        )
        assert "13:41:08" in disk and "file on disk" in disk


# --------------------------------------- the same table, through the real fake host


async def _docked_service(root: Path, name: str) -> OfficeHostService:
    """A real ``OfficeHostService`` with the fake backend and fake bridge, with
    ``name`` docked — the ``WORKBENCH_OFFICE_FAKE=1`` posture, which is what CI
    and the E2E journey run."""
    backend = FakeHostBackend()
    service = OfficeHostService(
        Workspace(root),
        EventBus(),
        backend,
        bridge=FakeDocumentBridge(backend),
        mode="on",
        fake=True,
        detector=lambda: False,
        poll_interval_s=60.0,
    )
    info = await service.open(name, RECT)
    assert info.state == "embedded", info
    return service


class TestThroughTheFakeHost:
    """No stub: the real service, the real bridge seam, the fake backend. If the
    wiring between ``ReconciliationCheck`` and ``OfficeHostService`` were wrong,
    every test above would still pass and this one would not."""

    async def test_a_dirty_docked_workbook_with_no_live_read_is_blocked(
        self, tmp_path: Path
    ) -> None:
        # `dirty` scripts Saved=False and `nolive` refuses the bulk value read —
        # the two halves of the blocked row, reachable with no Office at all.
        name = "budget-dirty-nolive.xlsx"
        make_workbook(tmp_path / name, {"B2": 1000.0})
        host = await _docked_service(tmp_path, name)
        _, result = await run_recon(tmp_path, cell_spec(name, "Budget!B2", 1000.0), host)

        assert result.risk == "blocked"
        assert "unsaved changes" in result.evidence[0].detail

    async def test_a_clean_docked_workbook_reads_live_values_as_numbers(
        self, tmp_path: Path
    ) -> None:
        # The fake mints Budget!B2 as 1000.0 — a float, not the string "1000".
        # Reconciling `"1000"` would be a `non-numeric cell` fail, which is
        # exactly why the live path cannot reuse `office_com.excel_grid`.
        name = "budget.xlsx"
        make_workbook(tmp_path / name, {"B2": 1.0})  # disk deliberately disagrees
        host = await _docked_service(tmp_path, name)
        service, result = await run_recon(tmp_path, cell_spec(name, "Budget!B2", 1000.0), host)

        assert result.risk == "pass", result.evidence[0].detail
        report = report_of(service, result)
        assert report.source.kind == "live"
        assert report.comparisons[0].actual == 1000.0

    async def test_an_office_write_dirties_the_workbook_and_the_gate_follows_it(
        self, tmp_path: Path
    ) -> None:
        """The journey, server-side: dock a workbook, edit a cell through the
        write seam an agent uses, and the gate judges the *edited* number and
        says the workbook is unsaved. On disk, the old number is still there."""
        name = "budget.xlsx"
        make_workbook(tmp_path / name, {"B2": 1000.0})
        host = await _docked_service(tmp_path, name)
        await host.write_document(name, content="9999", sheet="Budget", cell="B2")

        service, result = await run_recon(tmp_path, cell_spec(name, "Budget!B2", 1000.0), host)
        report = report_of(service, result)
        assert report.source.kind == "live"
        assert report.source.saved is False
        assert report.comparisons[0].actual == 9999.0
        assert result.risk == "high"
        # And the file on disk still says 1000.0 — which is what a gate without
        # this PR would have reconciled, and passed.
        disk = DiskWorkbookReader(tmp_path / name, datetime(2026, 8, 9, 14, 0))
        assert disk.cell_value("Sheet1", "B2") == 1000.0
        disk.close()

    async def test_a_calculating_workbook_is_never_a_pass(self, tmp_path: Path) -> None:
        name = "budget-calculating.xlsx"
        make_workbook(tmp_path / name, {"B2": 1000.0})
        host = await _docked_service(tmp_path, name)
        _, result = await run_recon(tmp_path, cell_spec(name, "Budget!B2", 1000.0), host)

        assert result.risk != "pass"
        assert any(item.outcome == "skipped" for item in result.evidence)

    async def test_a_workbook_nobody_docked_still_reads_disk(self, tmp_path: Path) -> None:
        """The no-regression case, and the one every existing CI run is: a host
        exists, this workbook is simply not in it."""
        make_workbook(tmp_path / "other.xlsx", {"B1": 1234.5})
        (tmp_path / "budget.xlsx").write_bytes(b"PK\x03\x04")
        host = await _docked_service(tmp_path, "budget.xlsx")
        service, result = await run_recon(tmp_path, cell_spec("other.xlsx", "B1", 1234.5), host)
        assert result.risk == "pass"
        assert report_of(service, result).source.kind == "file"
